from __future__ import annotations
from tracker_evaluation.evaluation_service import save_upload, get_tracker_interns, get_eval_scorecards, match_candidates, get_tracker_metrics, build_questions, suggest_score, finalize_evaluation
import os
import secrets
import string
from html import escape as escape_html

from pathlib import Path
from datetime import datetime
import json
import shutil

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles

from tracker_commands.executor import CommandExecutor
from tracker_commands.validator import CommandValidationError
from tracker_auth.user_service import UserService
from tracker_auth.permissions import can_execute, can_manage_users, can_view_logs, can_assign_role, can_modify_target, can_manage_admins
from tracker_audit.audit_service import AuditService
from tracker_tasks.task_service import TaskService
from tracker_audit.audit_db import init_db
from tracker_chat.chat_service import ChatService
from tracker_excel.renderer.parser import parse_workbook
from tracker_auth.jwt_service import create_session_token, decode_session_token, SESSION_TTL_SECONDS

BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
OUTPUT_DIR = BASE_DIR / "outputs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Intern Tracker Web UI", version="0.10")
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "web" / "static")), name="static")

# Only set the cookie's Secure flag once actually served over HTTPS (e.g.
# behind a TLS-terminating reverse proxy) - forcing it on in plain-HTTP dev
# would silently stop the browser from ever sending the session cookie back.
COOKIE_SECURE = os.getenv("COOKIE_SECURE", "false").strip().lower() == "true"


def _issue_csrf_cookie(res):
    # Double-submit CSRF cookie: deliberately NOT HttpOnly, since same-origin
    # JS (nav.js) needs to read it and echo it back as a header on every
    # mutating request - see the csrf_protection middleware below for the
    # actual check and why this is safe despite being JS-readable.
    token = secrets.token_hex(16)
    res.set_cookie('csrf_token', token, httponly=False, samesite='lax', secure=COOKIE_SECURE, max_age=SESSION_TTL_SECONDS)
    return res

# All pages here are server-rendered HTML with inline <script> blocks (no
# build step / bundler), so script-src and style-src need 'unsafe-inline'.
# Nothing in this app loads from a third-party origin, so default-src
# 'self' plus frame-ancestors 'none' still meaningfully blocks clickjacking
# and cross-site resource injection even though inline scripts stay allowed.
_CSP = (
    "default-src 'self'; "
    "script-src 'self' 'unsafe-inline'; "
    "style-src 'self' 'unsafe-inline'; "
    "img-src 'self' data:; "
    "frame-ancestors 'none'"
)


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Content-Security-Policy"] = _CSP
    # No explicit Cache-Control meant browsers fell back to heuristic
    # caching and could silently serve a stale page/JS/CSS file for a
    # while after a deploy - force revalidation on every load instead.
    # Originally scoped to /static/* only, which missed the page routes
    # themselves (/chat, /tasks, etc.) - a stale cached /chat page kept
    # showing old rendering logic even after the server-side fix landed.
    # Still fast for static assets (a 304 if unchanged, via the ETag
    # StaticFiles already sets) - this just guarantees the browser
    # actually asks, rather than assuming.
    response.headers["Cache-Control"] = "no-cache"
    return response


_CSRF_SAFE_METHODS = {"GET", "HEAD", "OPTIONS"}


@app.middleware("http")
async def csrf_protection(request: Request, call_next):
    # Double-submit cookie check for state-changing API requests. A
    # cross-site page can't read this origin's cookies (same-origin policy)
    # and can't attach a custom header via a plain HTML form submission, so
    # it can't forge a request carrying a matching X-CSRF-Token even though
    # the browser still attaches the session cookie automatically.
    # SameSite=Lax already blocks most of this; this covers the residual
    # gap (top-level GET-navigation CSRF, browsers with imperfect SameSite
    # support). Scoped to requests that already carry a session - an
    # anonymous request like POST /api/login has no pre-existing session
    # for a forged request to ride on in the first place.
    if request.method not in _CSRF_SAFE_METHODS and request.url.path.startswith("/api/"):
        if request.cookies.get("session_token"):
            csrf_cookie = request.cookies.get("csrf_token") or ""
            csrf_header = request.headers.get("x-csrf-token") or ""
            if not csrf_cookie or not secrets.compare_digest(csrf_cookie, csrf_header):
                return JSONResponse(status_code=403, content={"ok": False, "error": "Missing or invalid CSRF token"})
    return await call_next(request)


executor = CommandExecutor()
chat_service = ChatService()
init_db()
user_service = UserService()
audit_service = AuditService()
task_service = TaskService()


def safe_name(name: str) -> str:
    return Path(name).name.replace("/", "_").replace("\\", "_")


ALLOWED_UPLOAD_EXTENSIONS = {".xlsx", ".xls"}


def require_allowed_extension(filename: str):
    ext = Path(filename or "").suffix.lower()
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"Unsupported file type '{ext}'. Only .xlsx and .xls files are allowed.")


def user_upload_dir(user: dict) -> Path:
    d = UPLOAD_DIR / str(user["id"])
    d.mkdir(parents=True, exist_ok=True)
    return d


def user_output_dir(user: dict) -> Path:
    d = OUTPUT_DIR / str(user["id"])
    d.mkdir(parents=True, exist_ok=True)
    return d


def output_path(name: str | None, prefix: str, user: dict) -> str:
    folder = user_output_dir(user)
    if name:
        return str(folder / safe_name(name))
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(folder / f"{prefix}_{stamp}.xlsx")


def _within(path: Path, base: Path) -> bool:
    try:
        path.resolve().relative_to(base.resolve())
        return True
    except (OSError, ValueError):
        return False


def resolve_workbook(value: str, user: dict) -> str:
    if not value:
        raise HTTPException(status_code=400, detail="Workbook/source path is required")
    name = safe_name(value)
    candidates = [user_upload_dir(user) / name, user_output_dir(user) / name]
    if v65_is_admin_or_super(user):
        candidates += sorted(UPLOAD_DIR.glob(f"*/{name}")) + sorted(OUTPUT_DIR.glob(f"*/{name}"))
    for c in candidates:
        if Path(c).exists():
            return str(c)
    # A literal path is only trusted if it already resolves inside a
    # directory this user is allowed to read (their own upload/output dirs,
    # or any user's for Admin/Super Admin) - never a bare "does this path
    # exist anywhere on disk" check, which let any authenticated user read
    # any other user's files by guessing outputs/<user_id>/<filename>.xlsx.
    p = Path(value)
    allowed_bases = [user_upload_dir(user), user_output_dir(user)]
    if v65_is_admin_or_super(user):
        allowed_bases += [UPLOAD_DIR, OUTPUT_DIR]
    if p.exists() and any(_within(p, base) for base in allowed_bases):
        return str(p)
    # Fail closed: do not hand back an unvalidated path. A caller like
    # parse_workbook() would happily open it directly (relative to this
    # process's CWD), which is exactly how the old "return value" fallback
    # re-leaked cross-user files even after the checks above rejected them.
    raise HTTPException(status_code=404, detail="Workbook not found")


def _owned_files(folder: Path):
    return [{"name": p.name} for p in sorted(folder.glob("*.xlsx"))]


def _all_owned_files(base_dir: Path):
    out = []
    for sub in sorted(base_dir.iterdir()):
        if not sub.is_dir():
            continue
        owner = user_service.get_user_by_id(int(sub.name)) if sub.name.isdigit() else None
        owner_label = owner.get("email") if owner else "Unassigned (legacy)"
        for p in sorted(sub.glob("*.xlsx")):
            out.append({"name": p.name, "owner_id": sub.name, "owner": owner_label})
    return out




@app.get("/chat", response_class=HTMLResponse)
def chat_page(request: Request):
    if not current_user_from_request(request):
        return RedirectResponse('/login')
    return (BASE_DIR / "web" / "chat.html").read_text(encoding="utf-8")


@app.get("/workflow", response_class=HTMLResponse)
def workflow_page(request: Request):
    if not current_user_from_request(request):
        return RedirectResponse('/login')
    return (BASE_DIR / "web" / "workflow.html").read_text(encoding="utf-8")


@app.get("/forms", response_class=HTMLResponse)
def forms_page(request: Request):
    if not current_user_from_request(request):
        return RedirectResponse('/login')
    return (BASE_DIR / "web" / "forms.html").read_text(encoding="utf-8")



def current_user_from_request(request: Request):
    token = request.cookies.get('session_token')
    claims = decode_session_token(token)
    if not claims:
        return None
    user = user_service.get_user_by_id(int(claims['sub']))
    if not user or user.get('status') != 'Active':
        return None
    return user


def require_login(request: Request):
    user = current_user_from_request(request)
    if not user:
        raise HTTPException(status_code=401, detail='Login required')
    return user


def redirect_by_role(user):
    role = (user or {}).get('role', '')
    if role == 'Super Admin':
        return RedirectResponse('/dashboard/super-admin')
    if role == 'Admin':
        return RedirectResponse('/dashboard/admin')
    return RedirectResponse('/dashboard/user')



def v65_is_admin_or_super(user):
    role = (user or {}).get('role', '')
    return role in {'Super Admin', 'Admin'}

MAINTAINER_EMAIL = 'maintainer@example.com'

def is_maintainer(user):
    return (user or {}).get('email', '').strip().lower() == MAINTAINER_EMAIL

def v65_is_super(user):
    return (user or {}).get('role', '') == 'Super Admin'

@app.get('/login', response_class=HTMLResponse)
def login_page():
    return (BASE_DIR / 'web' / 'login.html').read_text(encoding='utf-8')

@app.post('/api/login')
def api_login(payload: dict):
    email = payload.get('email','')
    if user_service.is_locked_out(email):
        audit_service.log({'name':'Unknown','email':email}, interface='Auth', action='Login Blocked', status='Failed', summary='Too many failed attempts, temporarily locked out')
        return JSONResponse(status_code=429, content={'ok': False, 'error': 'Too many failed login attempts. Please try again in 15 minutes.'})
    user = user_service.authenticate(email, payload.get('password',''))
    if not user:
        user_service.record_failed_attempt(email)
        audit_service.log({'name':'Unknown','email':email}, interface='Auth', action='Login Failed', status='Failed', summary='Invalid login or inactive user')
        return JSONResponse(status_code=401, content={'ok': False, 'error': 'Invalid login or inactive user'})
    user_service.clear_attempts(email)
    audit_service.log(user, interface='Auth', action='Login', status='Success', summary='User logged in')
    token = create_session_token(user)
    public_user = {k: v for k, v in user.items() if k not in ('password', 'llm_api_key_encrypted')}
    public_user['is_maintainer'] = is_maintainer(user)
    res = JSONResponse({'ok': True, 'user': public_user})
    res.set_cookie('session_token', token, httponly=True, samesite='lax', secure=COOKIE_SECURE, max_age=SESSION_TTL_SECONDS)
    _issue_csrf_cookie(res)
    return res

@app.get('/logout')
def logout(request: Request):
    user = current_user_from_request(request)
    if user:
        user_service.logout(user['email'])
        audit_service.log(user, interface='Auth', action='Logout', status='Success', summary='User logged out')
    res = RedirectResponse('/login')
    res.delete_cookie('session_token')
    res.delete_cookie('csrf_token')
    return res

@app.get('/api/me')
def api_me(request: Request):
    user = current_user_from_request(request)
    if user:
        user = {**user, 'is_maintainer': is_maintainer(user)}
    return {'ok': bool(user), 'user': user}

@app.get('/users', response_class=HTMLResponse)
def users_page(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if not can_manage_users(user):
        return RedirectResponse('/chat')
    return (BASE_DIR / 'web' / 'users.html').read_text(encoding='utf-8')

@app.get('/logs', response_class=HTMLResponse)
def logs_page(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if not can_view_logs(user):
        return RedirectResponse('/chat')
    return (BASE_DIR / 'web' / 'logs.html').read_text(encoding='utf-8')

@app.get('/tasks', response_class=HTMLResponse)
def tasks_page(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    return (BASE_DIR / 'web' / 'tasks.html').read_text(encoding='utf-8')

@app.get('/ticket-queue', response_class=HTMLResponse)
def ticket_queue_page(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if not is_maintainer(user):
        return redirect_by_role(user)
    return (BASE_DIR / 'web' / 'ticket_queue.html').read_text(encoding='utf-8')

@app.get('/api/users')
def api_users(request: Request):
    user = require_login(request)
    if not can_manage_users(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin only'})
    return {'ok': True, 'users': user_service.list_users()}

@app.post('/api/users')
def api_create_user(request: Request, payload: dict):
    user = require_login(request)
    if not can_manage_users(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin only'})
    try:
        user_service.create_user(payload)
    except ValueError as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    audit_service.log(user, interface='Users', action='Create User', target_type='User', target_name=payload.get('email',''), status='Success')
    return {'ok': True}


@app.post('/api/users/reset-password')
def api_reset_password(request: Request, payload: dict):
    user = require_login(request)
    if not can_manage_users(user):
        audit_service.log(user, interface='Users', action='Reset Password', status='Blocked', summary='Admin only')
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin only'})
    target_id = int(payload.get('user_id'))
    new_password = payload.get('new_password', '')
    try:
        user_service.reset_password(target_id, new_password)
    except ValueError as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    audit_service.log(user, interface='Users', action='Reset Password', target_type='User', target_name=str(target_id), status='Success', summary='Admin reset user password')
    return {'ok': True}

@app.get('/api/logs')
def api_logs(request: Request, q: str = '', email: str = '', action: str = '', status: str = '', interface: str = ''):
    user = require_login(request)
    if not can_view_logs(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Not allowed'})
    filters = {'q': q, 'email': email, 'action': action, 'status': status, 'interface': interface}
    logs = audit_service.list_logs(filters=filters)
    return {'ok': True, 'logs': logs}

@app.get('/api/logs/export')
def api_logs_export(request: Request, q: str = '', email: str = '', action: str = '', status: str = '', interface: str = ''):
    user = require_login(request)
    if not can_view_logs(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Not allowed'})
    filters = {'q': q, 'email': email, 'action': action, 'status': status, 'interface': interface}
    path = audit_service.export_csv(filters=filters)
    audit_service.log(user, interface='Logs', action='Export Logs', status='Success', output_workbook=Path(path).name)
    return FileResponse(path, filename=Path(path).name)

# Rough estimate of manual time each automated action replaces - the
# system's own completion time is a few seconds, negligible against these,
# so "hours saved" collapses to just this manual-time estimate per action.
_HOURS_SAVED_MINUTES = {
    'add_intern_with_plan': 20, 'add_intern_basic': 20, 'add_intern': 20,
    'extend_intern_with_plan': 30, 'extend_intern': 30,
    'update_task_status': 2,
    'add_holiday': 5,
    'create_plan_from_draft': 30, 'create_plan': 30,
    'create_workbook': 10, 'render_workbook': 10,
}

@app.get('/api/dashboard/kpis')
def api_dashboard_kpis(request: Request):
    user = require_login(request)
    if not v65_is_admin_or_super(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    return {'ok': True, 'kpis': {
        'total_users': user_service.count_users(),
        'interns_added': audit_service.count_actions(['add_intern_with_plan', 'add_intern_basic', 'add_intern']),
        'workbooks_created': audit_service.count_actions(['create_workbook']),
        'plans_created': audit_service.count_actions(['create_plan_from_draft', 'create_plan']),
        'hours_saved': audit_service.hours_saved(_HOURS_SAVED_MINUTES),
        'top_plans': audit_service.top_targets(['create_plan_from_draft', 'create_plan']),
    }}

@app.get('/api/tasks')
def api_tasks(request: Request):
    user = require_login(request)
    if not is_maintainer(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Maintainer only'})
    return {'ok': True, 'tasks': task_service.list_tasks()}

@app.get('/api/tasks/pending-count')
def api_tasks_pending_count(request: Request):
    user = require_login(request)
    if not is_maintainer(user):
        return {'ok': True, 'count': 0}
    return {'ok': True, 'count': task_service.count_pending_tasks()}

@app.post('/api/tasks')
def api_create_task(request: Request, payload: dict):
    user = require_login(request)
    if not (payload.get('title') or '').strip():
        return JSONResponse(status_code=400, content={'ok': False, 'error': 'Title is required'})
    task_service.create_task(payload, user)
    audit_service.log(user, interface='Tasks', action='Create Task', target_type='Task', target_name=payload.get('title',''), status='Success')
    return {'ok': True}

@app.post('/api/tasks/request-api-key')
def api_request_api_key(request: Request, payload: dict):
    """Lets any logged-in user (not just Admin/Super Admin) raise this one
    specific kind of ticket, without opening up general ticket creation."""
    user = require_login(request)
    provider = (payload.get('provider') or '').strip()
    if provider not in {'Groq', 'Gemini', 'Other'}:
        provider = 'Other'
    task_service.create_task({
        'title': f"API Key Request - {user.get('name', '')}",
        'description': (payload.get('note') or '').strip() or 'Requesting an AI provider API key.',
        'category': 'API Key Request', 'priority': 'Medium', 'status': 'Pending',
        'assigned_to': '', 'due_date': '', 'remarks': '', 'ticket_type': provider,
    }, user)
    audit_service.log(user, interface='Tasks', action='Request API Key', status='Success')
    return {'ok': True}

_TASK_STATUS_VALUES = {'Pending', 'In Progress', 'Blocked', 'Completed', 'Cancelled'}

# Which status transitions are worth notifying the requester about, and what
# to tell them - Pending/In Progress are routine, but these three are either
# an endpoint or something that needs the requester's attention.
_STATUS_NOTIFY_MESSAGES = {
    'Completed': 'was resolved',
    'Blocked': 'is blocked - it may need more info from you',
    'Cancelled': 'was cancelled',
}

@app.post('/api/tasks/{task_id}/status')
def api_update_task_status(request: Request, task_id: int, payload: dict):
    user = require_login(request)
    task = task_service.get_task(task_id)
    if not task:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Ticket not found'})
    if not is_maintainer(user):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Maintainer only'})
    new_status = payload.get('status')
    if new_status not in _TASK_STATUS_VALUES:
        return JSONResponse(status_code=400, content={'ok': False, 'error': 'Invalid status'})
    old_status = task.get('status')
    merged = {**task, 'status': new_status}
    if 'resolution_note' in payload:
        merged['resolution_note'] = payload.get('resolution_note', '')
    if new_status in _STATUS_NOTIFY_MESSAGES and old_status != new_status:
        merged['creator_notified'] = 0
    task_service.update_task(task_id, merged)
    audit_service.log(user, interface='Tasks', action='Update Ticket Status', target_type='Task', target_name=task.get('title', ''), status='Success', summary=f'{old_status} -> {new_status}')
    return {'ok': True}

@app.get('/api/notifications')
def api_notifications(request: Request):
    user = require_login(request)
    items = task_service.list_unnotified_status_changes(user.get('name', ''))
    return {'ok': True, 'count': len(items), 'items': [{
        'id': t['id'], 'title': t.get('title') or '(untitled ticket)', 'description': t.get('description', ''),
        'category': t.get('category', ''), 'resolution_note': t.get('resolution_note', ''),
        'status': t.get('status', ''), 'status_text': _STATUS_NOTIFY_MESSAGES.get(t.get('status', ''), 'was updated'),
    } for t in items]}

@app.post('/api/notifications/mark-read')
def api_notifications_mark_read(request: Request):
    user = require_login(request)
    task_service.mark_notified(user.get('name', ''))
    return {'ok': True}

@app.post('/api/notifications/{task_id}/dismiss')
def api_notification_dismiss(request: Request, task_id: int):
    user = require_login(request)
    task = task_service.get_task(task_id)
    if not task or task.get('created_by') != user.get('name', ''):
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Not found'})
    task_service.mark_notified_one(task_id)
    return {'ok': True}

@app.get('/notifications', response_class=HTMLResponse)
def notifications_page(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    return (BASE_DIR / 'web' / 'notifications.html').read_text(encoding='utf-8')

@app.get("/")
def home(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    return redirect_by_role(user)


@app.get('/dashboard/super-admin', response_class=HTMLResponse)
def super_admin_dashboard(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if user.get('role') != 'Super Admin':
        return redirect_by_role(user)
    return (BASE_DIR / 'web' / 'super_admin_home.html').read_text(encoding='utf-8')


@app.get('/dashboard/admin', response_class=HTMLResponse)
def admin_dashboard(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if user.get('role') != 'Admin':
        return redirect_by_role(user)
    return (BASE_DIR / 'web' / 'admin_home.html').read_text(encoding='utf-8')


@app.get('/dashboard/user', response_class=HTMLResponse)
def user_dashboard(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if user.get('role') != 'User':
        return redirect_by_role(user)
    return (BASE_DIR / 'web' / 'user_home.html').read_text(encoding='utf-8')


@app.post("/api/upload")
def upload_workbook(request: Request, file: UploadFile = File(...)):
    user = require_login(request)
    require_allowed_extension(file.filename)
    filename = safe_name(file.filename or "workbook.xlsx")
    dst = user_upload_dir(user) / filename
    with dst.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    audit_service.log(user, interface='Forms', action='Upload Workbook', target_type='Workbook', target_name=filename, output_workbook=filename, status='Success')
    return {"ok": True, "filename": filename, "path": str(dst)}


@app.get("/api/files")
def list_files(request: Request):
    user = require_login(request)
    result = {"ok": True, "uploads": _owned_files(user_upload_dir(user)), "outputs": _owned_files(user_output_dir(user))}
    if v65_is_admin_or_super(user):
        result["all_uploads"] = _all_owned_files(UPLOAD_DIR)
        result["all_outputs"] = _all_owned_files(OUTPUT_DIR)
    return result


@app.post("/api/files/delete")
def delete_file(request: Request, payload: dict):
    user = require_login(request)
    folder = payload.get('folder')
    name = safe_name(payload.get('name', ''))
    owner_id = payload.get('owner_id')
    if folder not in {'uploads', 'outputs'} or not name:
        return JSONResponse(status_code=400, content={'ok': False, 'error': 'folder (uploads/outputs) and name are required'})
    base_dir = UPLOAD_DIR if folder == 'uploads' else OUTPUT_DIR
    if owner_id is not None and str(owner_id) != str(user['id']):
        if not v65_is_admin_or_super(user):
            return JSONResponse(status_code=403, content={'ok': False, 'error': 'You can only delete your own files'})
        target_dir = base_dir / safe_name(str(owner_id))
    else:
        target_dir = user_upload_dir(user) if folder == 'uploads' else user_output_dir(user)
    target = target_dir / name
    if target.resolve().parent != target_dir.resolve() or not target.exists() or not target.is_file():
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'File not found'})
    target.unlink()
    audit_service.log(user, interface='Forms', action='Delete Workbook', target_type='Workbook', target_name=name, status='Success')
    return {'ok': True}


@app.get("/download/{filename}")
def download(request: Request, filename: str):
    user = require_login(request)
    name = safe_name(filename)
    search_dirs = [user_output_dir(user), user_upload_dir(user)]
    if v65_is_admin_or_super(user):
        search_dirs += sorted(UPLOAD_DIR.glob("*")) + sorted(OUTPUT_DIR.glob("*"))
    for folder in search_dirs:
        p = Path(folder) / name
        if p.exists() and p.is_file():
            return FileResponse(str(p), filename=p.name)
    raise HTTPException(status_code=404, detail="File not found")


@app.post("/api/execute")
def execute_command(request: Request, payload: dict):
    user = require_login(request)
    try:
        cmd = payload.get("command")
        if not can_execute(user, cmd):
            audit_service.log(user, interface='Forms', action=cmd or 'Unknown', status='Blocked', summary='Permission denied')
            return JSONResponse(status_code=403, content={'ok': False, 'error': 'Permission denied'})
        args = payload.get("args") or {}
        # Normalize file fields for the web UI.
        if "source" in args:
            args["source"] = resolve_workbook(args["source"], user)
        if "workbook" in args:
            args["workbook"] = resolve_workbook(args["workbook"], user)
        if "output" in args:
            args["output"] = output_path(args.get("output"), cmd or "tracker", user)
        elif cmd not in {"summary"}:
            args["output"] = output_path(None, cmd or "tracker", user)
        result = executor.execute({"command": cmd, "args": args})
        response = result.public_dict()
        if result.output_path:
            response["download"] = f"/download/{Path(result.output_path).name}"
        audit_service.log(user, interface='Forms', action=cmd, target_name=args.get('intern') or args.get('name') or args.get('plan_name') or '', input_workbook=args.get('source') or args.get('workbook') or '', output_workbook=Path(result.output_path).name if result.output_path else '', status='Success' if result.ok else 'Failed', summary=response['message'])
        return response
    except CommandValidationError as e:
        audit_service.log(user, interface='Forms', action=payload.get('command',''), status='Failed', error_message=str(e))
        return JSONResponse(status_code=400, content={"ok": False, "error": str(e)})
    except Exception as e:
        audit_service.log(user, interface='Forms', action=payload.get('command',''), status='Failed', error_message=str(e))
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})


@app.post("/api/chat/message")
def chat_message(request: Request, payload: dict):
    user = require_login(request)
    text = payload.get('message', '')
    current_workbook = payload.get('current_workbook')
    if not text:
        return JSONResponse(status_code=400, content={'ok': False, 'error': 'message is required'})
    if current_workbook:
        current_workbook = resolve_workbook(current_workbook, user)
    return chat_service.message(text, current_workbook, user_service.get_user_llm_credentials(user['id']))

@app.post("/api/chat/update")
def chat_update(request: Request, payload: dict):
    user = require_login(request)
    return chat_service.update_draft(payload.get('draft_id'), payload.get('args') or {}, user_service.get_user_llm_credentials(user['id']))


@app.post("/api/chat/fill")
def chat_fill(request: Request, payload: dict):
    """Fill the active chat draft from a follow-up natural language message."""
    user = require_login(request)
    return chat_service.fill_from_text(payload.get('draft_id'), payload.get('message', ''), user_service.get_user_llm_credentials(user['id']))

@app.post("/api/chat/approve")
def chat_approve(request: Request, payload: dict):
    """Approve and execute a chat draft.

    Important: chat drafts are created in the browser using the Current Workbook label.
    That label may be just a filename such as Rendered_Extended.xlsx. Before executing,
    normalize source/workbook/output paths exactly like /api/execute does, otherwise
    openpyxl may look in the project root and fail when the file is actually in outputs/.
    """
    user = require_login(request)
    draft_id = payload.get('draft_id')
    draft = getattr(chat_service, 'drafts', {}).get(draft_id)
    if not draft:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Draft not found'})
    if not can_execute(user, draft.command):
        audit_service.log(user, interface='Chat', action=getattr(draft, 'command', 'chat_approve'), status='Blocked', summary='Permission denied')
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Permission denied'})
    try:
        args = draft.args
        if 'source' in args:
            args['source'] = resolve_workbook(args['source'], user)
        if 'workbook' in args:
            args['workbook'] = resolve_workbook(args['workbook'], user)
        if 'spec' in args:
            # add_intern's JSON spec file is looked up the same sandboxed
            # way as a workbook (own upload/output dir, or any user's for
            # Admin/Super Admin) - never a raw filesystem path, which would
            # let chat text or a direct /api/chat/update call point this at
            # any file the server process can read.
            args['spec'] = resolve_workbook(args['spec'], user)
        if 'output' in args:
            args['output'] = output_path(args.get('output'), draft.command or 'chat', user)
        elif draft.command not in {'summary'}:
            args['output'] = output_path(None, draft.command or 'chat', user)
        audit_service.log(user, interface='Chat', action=getattr(draft, 'command', 'chat_approve'), approval_status='Approved', status='Started', summary='Chat proposal approved')
        result = chat_service.approve(draft_id)
        audit_service.log(user, interface='Chat', action=getattr(draft, 'command', 'chat_approve'), target_name=(getattr(draft, 'args', {}) or {}).get('intern') or (getattr(draft, 'args', {}) or {}).get('name') or (getattr(draft, 'args', {}) or {}).get('plan_name') or '', input_workbook=(getattr(draft, 'args', {}) or {}).get('source') or (getattr(draft, 'args', {}) or {}).get('workbook') or '', output_workbook=Path(result.get('output_path','')).name if result.get('output_path') else '', approval_status='Approved', status='Success' if result.get('ok') else 'Failed', summary=result.get('message',''), error_message=result.get('error',''))
        if result.get('output_path'):
            result['download'] = f"/download/{Path(result['output_path']).name}"
        # Opt-in preference (Profile page): keep only the latest version of a
        # workbook instead of a new file per action forever. Only ever
        # deletes a file inside this user's own upload/output directory -
        # the ownership check is what keeps this safe for an Admin/Super
        # Admin who might have another user's or a legacy file selected.
        if result.get('ok') and user.get('auto_cleanup_versions'):
            try:
                old_source = Path(args.get('source') or '')
                # Use the already-computed full path, not result['output_path']
                # (now just a filename - see CommandResult.public_dict()).
                new_output = Path(args.get('output') or '')
                own_dirs = {user_upload_dir(user).resolve(), user_output_dir(user).resolve()}
                if old_source.exists() and old_source.resolve().parent in own_dirs and old_source.resolve() != new_output.resolve():
                    old_source.unlink()
            except Exception:
                pass
        return result
    except Exception as e:
        return JSONResponse(status_code=500, content={'ok': False, 'error': str(e)})

@app.post("/api/chat/cancel")
def chat_cancel(request: Request, payload: dict):
    require_login(request)
    return chat_service.cancel(payload.get('draft_id'))


@app.post("/api/chat/manual")
def chat_manual(request: Request, payload: dict):
    """Create a draft directly from structured Forms UI input, bypassing
    all text-parsing. Same draft lifecycle as /api/chat/message from here
    on - the returned draft_id works with update/approve/cancel unchanged.
    """
    user = require_login(request)
    command = payload.get('command', '')
    args = payload.get('args') or {}
    current_workbook = payload.get('current_workbook')
    if current_workbook:
        current_workbook = resolve_workbook(current_workbook, user)
    return chat_service.create_manual_draft(command, args, current_workbook, user_service.get_user_llm_credentials(user['id']))


@app.get("/api/workbook/interns")
def workbook_interns(request: Request, source: str = ''):
    """List intern names in a workbook, for Forms UI dropdowns."""
    user = require_login(request)
    if not source:
        return {'ok': True, 'interns': []}
    try:
        path = resolve_workbook(source, user)
        data = parse_workbook(path)
        return {'ok': True, 'interns': [i.name for i in data.interns]}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})


@app.get("/api/workbook/plans")
def workbook_plans(request: Request, source: str = ''):
    """List plan names in a workbook, for Forms UI dropdowns.

    Plan sheet titles are stored as "Plan — <name>" (see plan_service.py's
    create_plan_from_draft); strip that prefix so the dropdown shows just
    the name, and so the value submitted back matches what _find_plan's
    substring match against .title expects.
    """
    user = require_login(request)
    if not source:
        return {'ok': True, 'plans': []}
    try:
        path = resolve_workbook(source, user)
        data = parse_workbook(path)
        names = []
        for p in data.plans:
            title = (p.title or '').strip()
            name = title.split('—', 1)[1].strip() if '—' in title else (title or p.sheet_name)
            if name:
                names.append(name)
        return {'ok': True, 'plans': names}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})


@app.get("/api/workbook/tasks")
def workbook_tasks(request: Request, source: str = '', intern: str = ''):
    """List an intern's tasks (date/theme/status), for the Forms UI task picker."""
    user = require_login(request)
    if not source or not intern:
        return {'ok': True, 'tasks': []}
    try:
        path = resolve_workbook(source, user)
        data = parse_workbook(path)
        match = next((i for i in data.interns if i.name.strip().lower() == intern.strip().lower()), None)
        if not match:
            return {'ok': True, 'tasks': []}
        # Substring match, not exact - actual headers are things like
        # "Status (Pending/In Progress/Completed)", not the bare word.
        headers = [str(h or '').strip().lower() for h in (match.task_headers or [])]
        def col(*names):
            for i, h in enumerate(headers):
                if any(n in h for n in names):
                    return i
            return None
        date_i, theme_i, status_i = col('date'), col('theme', 'task'), col('status')
        tasks = []
        for row in match.tasks:
            date_val = row[date_i] if date_i is not None and date_i < len(row) else None
            if date_val is None:
                continue
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)
            theme_val = row[theme_i] if theme_i is not None and theme_i < len(row) else ''
            status_val = row[status_i] if status_i is not None and status_i < len(row) else ''
            tasks.append({'date': date_str, 'theme': str(theme_val or ''), 'status': str(status_val or '')})
        return {'ok': True, 'tasks': tasks}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})


# v0.61 approval-role governance routes
@app.get('/signup', response_class=HTMLResponse)
def signup_page():
    return (BASE_DIR / 'web' / 'signup.html').read_text(encoding='utf-8')

@app.get('/pending', response_class=HTMLResponse)
def pending_page():
    return (BASE_DIR / 'web' / 'pending.html').read_text(encoding='utf-8')

@app.get('/profile', response_class=HTMLResponse)
def profile_page(request: Request):
    if not current_user_from_request(request):
        return RedirectResponse('/login')
    return (BASE_DIR / 'web' / 'profile.html').read_text(encoding='utf-8')

@app.post('/api/signup')
def api_signup(payload: dict):
    try:
        user_service.signup(payload)
        audit_service.log({'name': payload.get('name',''), 'email': payload.get('email','')}, interface='Auth', action='Signup Request', target_type='User', target_name=payload.get('email',''), status='Pending', summary='User requested access')
        return {'ok': True, 'message': 'Signup request submitted. Please wait for admin approval.'}
    except Exception as e:
        err = str(e) or 'Signup failed.'
        audit_service.log({'name': payload.get('name',''), 'email': payload.get('email','')}, interface='Auth', action='Signup Request Failed', target_type='User', target_name=payload.get('email',''), status='Failed', error_message=err)
        return JSONResponse(status_code=400, content={'ok': False, 'error': err})

@app.post('/api/users/approve')
def api_approve_user(request: Request, payload: dict):
    actor = require_login(request)
    if not can_manage_users(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    target_id = int(payload.get('user_id'))
    role = payload.get('role') or 'User'
    target = user_service.get_user_by_id(target_id)
    if not target:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'User not found'})
    if not can_assign_role(actor, role):
        audit_service.log(actor, interface='Users', action='Approve User', target_type='User', target_name=target.get('email',''), status='Blocked', summary=f'Role assignment blocked: {role}')
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'You cannot approve this role'})
    user_service.approve_user(target_id, role)
    audit_service.log(actor, interface='Users', action='Approve User', target_type='User', target_name=target.get('email',''), status='Success', summary=f'Approved as {role}')
    return {'ok': True}

@app.post('/api/users/reject')
def api_reject_user(request: Request, payload: dict):
    actor = require_login(request)
    if not can_manage_users(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    target_id = int(payload.get('user_id'))
    target = user_service.get_user_by_id(target_id)
    if not target:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'User not found'})
    if target.get('role') in {'Admin', 'Super Admin'} and not can_manage_admins(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Only Super Admin can reject/admin-manage admins'})
    user_service.reject_user(target_id)
    audit_service.log(actor, interface='Users', action='Reject User', target_type='User', target_name=target.get('email',''), status='Success')
    return {'ok': True}

@app.post('/api/users/deactivate')
def api_deactivate_user(request: Request, payload: dict):
    actor = require_login(request)
    if not can_manage_users(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    target_id = int(payload.get('user_id'))
    target = user_service.get_user_by_id(target_id)
    if not target:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'User not found'})
    if target.get('role') == 'Super Admin':
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Super Admin cannot be deactivated here'})
    if not can_modify_target(actor, target):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Only Super Admin can manage Admins'})
    user_service.deactivate_user(target_id)
    audit_service.log(actor, interface='Users', action='Deactivate User', target_type='User', target_name=target.get('email',''), status='Success')
    return {'ok': True}

@app.post('/api/users/role')
def api_change_user_role(request: Request, payload: dict):
    actor = require_login(request)
    if not can_manage_admins(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Only Super Admin can change Admin/User roles'})
    target_id = int(payload.get('user_id'))
    role = payload.get('role') or 'User'
    target = user_service.get_user_by_id(target_id)
    if not target:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'User not found'})
    if target.get('role') == 'Super Admin':
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Cannot change Super Admin role'})
    user_service.update_role(target_id, role)
    audit_service.log(actor, interface='Users', action='Change Role', target_type='User', target_name=target.get('email',''), status='Success', summary=f'Role changed to {role}')
    return {'ok': True}

@app.post('/api/profile')
def api_update_profile(request: Request, payload: dict):
    actor = require_login(request)
    old_email = actor.get('email')
    try:
        user_service.update_profile(old_email, payload)
    except ValueError as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    audit_service.log(actor, interface='Profile', action='Update Profile', status='Success', summary='User updated own profile')
    updated = user_service.get_user_by_id(actor['id'])
    res = JSONResponse({'ok': True})
    if updated:
        res.set_cookie('session_token', create_session_token(updated), httponly=True, samesite='lax', secure=COOKIE_SECURE, max_age=SESSION_TTL_SECONDS)
        _issue_csrf_cookie(res)
    return res


@app.get('/api/profile/llm-key-preview')
def api_llm_key_preview(request: Request):
    """Masked preview of the caller's own stored API key - never returns the
    raw or encrypted value, only whether one is set and its last 4 chars."""
    user = require_login(request)
    creds = user_service.get_user_llm_credentials(user['id'])
    masked = ''
    if creds.get('llm_api_key_encrypted'):
        try:
            from tracker_auth.key_crypto import decrypt_api_key, mask_api_key
            masked = mask_api_key(decrypt_api_key(creds['llm_api_key_encrypted']))
        except Exception:
            masked = ''
    return {'ok': True, 'llm_provider': creds.get('llm_provider') or '', 'llm_model': creds.get('llm_model') or '', 'llm_base_url': creds.get('llm_base_url') or '', 'has_key': bool(masked), 'masked': masked}


@app.post('/api/users/reactivate')
def api_reactivate_user(request: Request, payload: dict):
    actor = require_login(request)
    if not can_manage_users(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    target_id = int(payload.get('user_id'))
    target = user_service.get_user_by_id(target_id) if hasattr(user_service, 'get_user_by_id') else None
    if not target:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'User not found'})
    if target.get('role') == 'Super Admin':
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Super Admin cannot be reactivated here'})
    # Admin can reactivate normal Users only. Super Admin can reactivate Admins and Users.
    if actor.get('role') == 'Admin' and target.get('role') != 'User':
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admins can reactivate Users only'})
    if actor.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    conn = user_service.__class__.__dict__.get('get_conn', None)
    # Use audit DB connection directly for compatibility with monkey-patched UserService.
    from tracker_audit.audit_db import get_conn
    db = get_conn()
    db.execute('UPDATE users SET status=? WHERE id=?', ('Active', target_id))
    db.commit()
    db.close()
    audit_service.log(actor, interface='Users', action='Reactivate User', target_type='User', target_name=target.get('email',''), status='Success')
    return {'ok': True}


# v0.67 password reset request routes
from tracker_audit.audit_db import get_conn
try:
    from tracker_auth.passwords import hash_password
except Exception:
    import base64, hashlib, os
    def hash_password(password: str) -> str:
        salt = os.urandom(16)
        iterations = 260000
        dk = hashlib.pbkdf2_hmac('sha256', (password or '').encode('utf-8'), salt, iterations)
        return f'pbkdf2_sha256${iterations}${base64.b64encode(salt).decode()}${base64.b64encode(dk).decode()}'


def v67_init_password_reset_table():
    conn = get_conn()
    conn.execute('''
    CREATE TABLE IF NOT EXISTS password_reset_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT NOT NULL,
        user_id INTEGER,
        status TEXT DEFAULT 'Pending',
        requested_at TEXT NOT NULL,
        processed_by TEXT DEFAULT '',
        processed_at TEXT DEFAULT '',
        remarks TEXT DEFAULT ''
    )
    ''')
    conn.commit()
    conn.close()


def v67_now():
    from datetime import datetime
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def v67_generate_temp_password():
    alphabet = string.ascii_letters + string.digits
    token = ''.join(secrets.choice(alphabet) for _ in range(8))
    return 'Temp@' + token


def v67_user_by_email(email: str):
    conn = get_conn()
    row = conn.execute('SELECT id,name,email,department,role,status FROM users WHERE lower(email)=lower(?)', (email,)).fetchone()
    conn.close()
    return dict(row) if row else None


v67_init_password_reset_table()

@app.get('/forgot-password', response_class=HTMLResponse)
def forgot_password_page():
    return (BASE_DIR / 'web' / 'forgot_password.html').read_text(encoding='utf-8')

@app.post('/api/password-reset/request')
def api_password_reset_request(payload: dict):
    email = (payload.get('email') or '').strip().lower()
    remarks = (payload.get('remarks') or '').strip()
    if not email or '@' not in email:
        return JSONResponse(status_code=400, content={'ok': False, 'error': 'Valid email is required.'})
    target = v67_user_by_email(email)
    if not target:
        # Avoid leaking too much, but still log for admin visibility.
        audit_service.log({'name':'Unknown','email':email}, interface='Auth', action='Password Reset Request', target_type='User', target_name=email, status='Pending', summary='Password reset requested for email not currently active in users table')
        return {'ok': True, 'message': 'If this email exists, a reset request has been submitted.'}
    conn = get_conn()
    existing = conn.execute("SELECT id FROM password_reset_requests WHERE lower(email)=lower(?) AND status='Pending'", (email,)).fetchone()
    if existing:
        conn.close()
        return {'ok': True, 'message': 'A pending reset request already exists for this email.'}
    conn.execute('''INSERT INTO password_reset_requests(email,user_id,status,requested_at,remarks)
                    VALUES(?,?,?,?,?)''', (email, target.get('id'), 'Pending', v67_now(), remarks))
    conn.commit()
    conn.close()
    audit_service.log({'name':target.get('name',''), 'email':email}, interface='Auth', action='Password Reset Request', target_type='User', target_name=email, status='Pending', summary='User requested password reset')
    return {'ok': True, 'message': 'Password reset request submitted. Please wait for Admin/Super Admin approval.'}

@app.get('/api/password-reset/requests')
def api_password_reset_requests(request: Request):
    actor = require_login(request)
    if not v65_is_admin_or_super(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    conn = get_conn()
    rows = conn.execute('''SELECT r.*, u.name, u.role, u.status AS user_status
                           FROM password_reset_requests r
                           LEFT JOIN users u ON u.id=r.user_id
                           ORDER BY r.id DESC LIMIT 200''').fetchall()
    conn.close()
    return {'ok': True, 'requests': [dict(x) for x in rows]}

@app.post('/api/password-reset/complete')
def api_password_reset_complete(request: Request, payload: dict):
    actor = require_login(request)
    if not v65_is_admin_or_super(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    req_id = int(payload.get('request_id'))
    conn = get_conn()
    req = conn.execute('''SELECT r.*, u.role, u.email AS user_email, u.name AS user_name
                          FROM password_reset_requests r
                          LEFT JOIN users u ON u.id=r.user_id
                          WHERE r.id=?''', (req_id,)).fetchone()
    if not req:
        conn.close()
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Request not found'})
    req = dict(req)
    if req.get('status') != 'Pending':
        conn.close()
        return JSONResponse(status_code=400, content={'ok': False, 'error': 'Request is not pending'})
    target_role = req.get('role') or ''
    if actor.get('role') == 'Admin' and target_role != 'User':
        conn.close()
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admins can reset normal User passwords only'})
    if target_role == 'Super Admin':
        conn.close()
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Super Admin password must be changed from profile or handled directly by Super Admin'})
    temp_password = v67_generate_temp_password()
    conn.execute('UPDATE users SET password=? WHERE id=?', (hash_password(temp_password), req.get('user_id')))
    conn.execute('UPDATE password_reset_requests SET status=?, processed_by=?, processed_at=? WHERE id=?', ('Completed', actor.get('email',''), v67_now(), req_id))
    conn.commit()
    conn.close()
    audit_service.log(actor, interface='Users', action='Password Reset Completed', target_type='User', target_name=req.get('email',''), status='Success', summary='Temporary password generated. Password value not stored in logs.')
    # Return temp password once to admin/super admin so they can share via approved channel.
    return {'ok': True, 'temporary_password': temp_password, 'message': 'Temporary password generated. Share it with the user through an approved channel. User should change it after login.'}

@app.post('/api/password-reset/reject')
def api_password_reset_reject(request: Request, payload: dict):
    actor = require_login(request)
    if not v65_is_admin_or_super(actor):
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    req_id = int(payload.get('request_id'))
    conn = get_conn()
    req = conn.execute('SELECT * FROM password_reset_requests WHERE id=?', (req_id,)).fetchone()
    if not req:
        conn.close()
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Request not found'})
    conn.execute('UPDATE password_reset_requests SET status=?, processed_by=?, processed_at=? WHERE id=?', ('Rejected', actor.get('email',''), v67_now(), req_id))
    conn.commit()
    conn.close()
    audit_service.log(actor, interface='Users', action='Password Reset Rejected', target_type='User', target_name=dict(req).get('email',''), status='Success')
    return {'ok': True}


EVAL_SESSIONS = {}


@app.get('/evaluation', response_class=HTMLResponse)
def evaluation_page(request: Request):
    # v0.81 evaluation route protection
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if user.get('role') not in {'Admin', 'Super Admin'}:
        return RedirectResponse('/chat')
    return (BASE_DIR / 'web' / 'evaluation.html').read_text(encoding='utf-8')

@app.post('/api/evaluation/upload')
def api_evaluation_upload(request: Request, tracker: UploadFile = File(...), evaluation: UploadFile = File(...)):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    for f in (tracker, evaluation):
        ext = Path(f.filename or '').suffix.lower()
        if ext not in ALLOWED_UPLOAD_EXTENSIONS:
            return JSONResponse(status_code=400, content={'ok': False, 'error': f"Unsupported file type '{ext}'. Only .xlsx and .xls files are allowed."})
    try:
        tracker_path = save_upload(tracker, 'eval_tracker')
        eval_path = save_upload(evaluation, 'eval_framework')
        tracker_interns = get_tracker_interns(str(tracker_path))
        eval_cards = get_eval_scorecards(str(eval_path))
        session_id = __import__('uuid').uuid4().hex
        EVAL_SESSIONS[session_id] = {'tracker': str(tracker_path), 'evaluation': str(eval_path), 'user': user.get('email')}
        audit_service.log(user, interface='Evaluation', action='Upload Evaluation Files', status='Success', summary=f'{len(tracker_interns)} tracker interns, {len(eval_cards)} scorecards')
        return {'ok': True, 'session_id': session_id, 'tracker_interns': tracker_interns, 'eval_cards': eval_cards}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})

@app.post('/api/evaluation/questions')
def api_evaluation_questions(request: Request, payload: dict):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    sess = EVAL_SESSIONS.get(payload.get('session_id'))
    if not sess:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Evaluation session not found'})
    metrics = get_tracker_metrics(sess['tracker'], payload.get('intern_name',''), payload.get('evaluation_date'), payload.get('basis', 'as_of'))
    questions = build_questions()
    return {'ok': True, 'metrics': metrics, 'questions': questions}

@app.post('/api/evaluation/suggest')
def api_evaluation_suggest(request: Request, payload: dict):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    try:
        result = suggest_score(payload.get('criterion',''), payload.get('answer',''), payload.get('metrics') or {})
        return {'ok': True, **result}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})

@app.post('/api/evaluation/finalize')
def api_evaluation_finalize(request: Request, payload: dict):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    sess = EVAL_SESSIONS.get(payload.get('session_id'))
    if not sess:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Evaluation session not found'})
    try:
        out, missing = finalize_evaluation(sess['evaluation'], payload.get('eval_sheet'), payload.get('metrics') or {}, payload.get('scores') or {}, {}, payload.get('strengths',''), payload.get('development',''), payload.get('remark',''))
        summary = 'Some fields could not be written - labels not found: ' + ', '.join(missing) if missing else ''
        audit_service.log(user, interface='Evaluation', action='Finalize Evaluation', target_type='Intern', target_name=payload.get('intern_name',''), output_workbook=Path(out).name, status='Success', summary=summary)
        return {'ok': True, 'output_path': Path(out).name, 'download': '/evaluation/download?file=' + Path(out).name, 'warnings': missing}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})


@app.get('/evaluation/download')
def evaluation_download(request: Request, file: str):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    # Only allow downloading files from outputs by filename, not arbitrary paths.
    safe_name = Path(file).name
    target = BASE_DIR / 'outputs' / safe_name
    if not target.exists() or not target.is_file():
        return JSONResponse(status_code=404, content={'ok': False, 'error': f'Evaluation output not found: {safe_name}'})
    return FileResponse(str(target), filename=safe_name, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# v0.92 read-only intern progress summary endpoint
@app.post('/api/readonly/intern-summary')
def api_v92_readonly_intern_summary(request: Request, payload: dict):
    user = require_login(request)
    try:
        from pathlib import Path
        from difflib import SequenceMatcher
        from datetime import datetime, date
        from openpyxl import load_workbook
        import re

        def norm(v):
            return re.sub(r'\s+', ' ', str(v or '').strip())

        def norm_name(v):
            return re.sub(r'[^a-z0-9]+', ' ', str(v or '').lower()).strip()

        def status_done(v):
            return str(v or '').strip().lower() in {'completed', 'complete', 'done'}

        def parse_date(v):
            if hasattr(v, 'date'):
                return v.date()
            return None

        def pct(a, b):
            return (a / b) if b else 0

        def extract_requested_name(prompt):
            p = norm(prompt)
            p2 = re.sub(r'(?i)^how\s+is\s+', '', p)
            p2 = re.sub(r'(?i)\s+doing\??$', '', p2)
            p2 = re.sub(r'(?i)^show\s+', '', p2)
            p2 = re.sub(r'(?i)\s+progress\??$', '', p2)
            p2 = re.sub(r'(?i)^summar(?:y|ize|ise)\s+', '', p2)
            return norm(p2) or p

        def resolve_workbook(value):
            raw = norm(value)
            own_output = user_output_dir(user)
            own_upload = user_upload_dir(user)
            candidates = []
            if raw:
                cleaned = raw.replace('outputs /', 'outputs/').replace('uploads /', 'uploads/')
                literal = Path(cleaned)
                # Same rule as the module-level resolve_workbook(): only
                # trust the literal path if it's already inside a
                # directory this user may read - otherwise any
                # authenticated user could pass outputs/<other id>/<file>
                # and read another user's workbook straight through.
                allowed_bases = [own_upload, own_output]
                if v65_is_admin_or_super(user):
                    allowed_bases += [UPLOAD_DIR, OUTPUT_DIR]
                if any(_within(literal, base) for base in allowed_bases):
                    candidates.append(literal)
                candidates.append(own_output / literal.name)
                candidates.append(own_upload / literal.name)
                if v65_is_admin_or_super(user):
                    candidates += sorted(OUTPUT_DIR.glob(f"*/{literal.name}")) + sorted(UPLOAD_DIR.glob(f"*/{literal.name}"))
            for c in candidates:
                if Path(c).exists() and Path(c).is_file():
                    return Path(c)
            # Fallback to the user's own most recently modified workbook.
            outs = sorted(own_output.glob('*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
            if outs:
                return outs[0]
            ups = sorted(own_upload.glob('*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True)
            if ups:
                return ups[0]
            return None

        def discover_intern_sheets(wb):
            items = []
            for ws in wb.worksheets:
                title = str(ws.title or '')
                a1 = str(ws['A1'].value or '')
                if 'Intern Tracker' in a1:
                    name = a1.split('—', 1)[-1].split('(', 1)[0].strip() if '—' in a1 else title
                    items.append((name, title))
            return items

        def find_best_intern(wb, requested):
            sheets = discover_intern_sheets(wb)
            req = norm_name(requested)
            if not sheets:
                return None, None, []
            scored = []
            req_tokens = set(req.split())
            for name, sheet in sheets:
                nn = norm_name(name)
                ns = norm_name(sheet)
                tokens = set(nn.split()) | set(ns.split())
                overlap = len(req_tokens & tokens) / max(len(req_tokens), 1)
                ratio = max(SequenceMatcher(None, req, nn).ratio(), SequenceMatcher(None, req, ns).ratio())
                score = max(overlap, ratio)
                scored.append((score, name, sheet))
            scored.sort(reverse=True)
            best = scored[0]
            return best[1], best[2], [{'name': n, 'sheet': sh, 'score': round(sc, 3)} for sc, n, sh in scored[:5]]

        def find_row(ws, label):
            target = str(label).strip().lower()
            for row in ws.iter_rows():
                for cell in row:
                    if str(cell.value or '').strip().lower().startswith(target):
                        return cell.row
            return None

        def parse_daily(ws):
            row = find_row(ws, 'DAILY TASKS')
            if not row:
                return []
            header = row + 1
            headers = {str(ws.cell(header, c).value or '').strip().lower(): c for c in range(1, ws.max_column + 1)}
            c_date = headers.get('date')
            c_week = headers.get('week') or headers.get('week #')
            c_theme = headers.get('theme')
            c_task = headers.get('task description') or headers.get('task')
            c_status = next((c for h, c in headers.items() if 'status' in h), None)
            tasks = []
            if not c_status:
                return tasks
            for r in range(header + 1, ws.max_row + 1):
                first = str(ws.cell(r, 1).value or '').strip().lower()
                if first in {'weekly updates', 'small projects / tasks', 'small projects / tasks  (weekly projects)', 'main project', 'real-world scenario'}:
                    break
                status = ws.cell(r, c_status).value
                if status in (None, ''):
                    continue
                tasks.append({
                    'date': parse_date(ws.cell(r, c_date).value) if c_date else None,
                    'week': ws.cell(r, c_week).value if c_week else '',
                    'theme': ws.cell(r, c_theme).value if c_theme else '',
                    'task': ws.cell(r, c_task).value if c_task else '',
                    'status': status,
                })
            return tasks

        def parse_projects(ws):
            row = None
            for r in range(1, ws.max_row + 1):
                if str(ws.cell(r, 1).value or '').strip().lower().startswith('small projects / tasks'):
                    row = r
                    break
            if not row:
                return []
            header = row + 1
            headers = {str(ws.cell(header, c).value or '').strip().lower(): c for c in range(1, ws.max_column + 1)}
            c_title = headers.get('title')
            c_status = next((c for h, c in headers.items() if 'status' in h), None)
            projects = []
            if not c_status:
                return projects
            for r in range(header + 1, ws.max_row + 1):
                status = ws.cell(r, c_status).value
                if status in (None, ''):
                    continue
                projects.append({'title': ws.cell(r, c_title).value if c_title else '', 'status': status})
            return projects

        prompt = payload.get('prompt', '')
        wb_path = resolve_workbook(payload.get('workbook', ''))
        if not wb_path:
            return JSONResponse(status_code=400, content={'ok': False, 'error': 'No workbook found. Please select or upload a workbook first.'})

        wb = load_workbook(wb_path, data_only=True)
        requested = extract_requested_name(prompt)
        intern_name, sheet_name, candidates = find_best_intern(wb, requested)
        if not sheet_name:
            return JSONResponse(status_code=404, content={'ok': False, 'error': f'No intern tracker sheets found in {wb_path.name}.'})
        ws = wb[sheet_name]
        tasks = parse_daily(ws)
        projects = parse_projects(ws)

        total = len(tasks)
        completed = sum(1 for t in tasks if status_done(t['status']))
        in_progress = sum(1 for t in tasks if str(t['status']).strip().lower() == 'in progress')
        pending = total - completed - in_progress
        completion = pct(completed, total)
        ptotal = len(projects)
        pdone = sum(1 for p in projects if status_done(p['status']))
        pcompletion = pct(pdone, ptotal)

        completed_weeks = []
        pending_weeks = []
        for t in tasks:
            wk = t.get('week')
            if status_done(t['status']):
                completed_weeks.append(wk)
            elif wk not in ('', None):
                pending_weeks.append(wk)
        current_week = pending_weeks[0] if pending_weeks else ('Done' if total and completed == total else '')

        completed_themes = []
        pending_themes = []
        for t in tasks:
            theme = norm(t.get('theme'))
            if not theme:
                continue
            if status_done(t['status']) and theme not in completed_themes:
                completed_themes.append(theme)
            if not status_done(t['status']) and theme not in pending_themes:
                pending_themes.append(theme)

        status_label = 'On Track' if completion >= 0.75 else ('Developing' if completion >= 0.45 else 'Needs Support')
        # intern_name/workbook filename/week label/theme text all ultimately
        # come from the uploaded Excel file's own content - escape before
        # embedding in HTML sent to the browser, since this response gets
        # rendered via innerHTML client-side with no escaping of its own.
        safe_intern_name = escape_html(str(intern_name))
        safe_wb_name = escape_html(wb_path.name)
        safe_current_week = escape_html(str(current_week))
        safe_completed_areas = ', '.join(escape_html(str(t)) for t in completed_themes[:4]) or 'Not available from tracker.'
        safe_pending_areas = ', '.join(escape_html(str(t)) for t in pending_themes[:4]) or 'No pending areas found.'
        html = f"""
<h3>{safe_intern_name} - Progress Summary</h3>
<ul>
  <li><b>Workbook:</b> {safe_wb_name}</li>
  <li><b>Daily tasks:</b> {completed}/{total} completed ({completion:.0%})</li>
  <li><b>Weekly projects:</b> {pdone}/{ptotal} completed ({pcompletion:.0%})</li>
  <li><b>In progress:</b> {in_progress}</li>
  <li><b>Pending:</b> {pending}</li>
  <li><b>Current/next week:</b> {safe_current_week}</li>
  <li><b>Status:</b> {status_label}</li>
</ul>
<p><b>Completed areas:</b> {safe_completed_areas}</p>
<p><b>Pending/upcoming areas:</b> {safe_pending_areas}</p>
<p><b>Suggested manager action:</b> Review pending tasks/projects and ask for blockers if progress is below expected pace.</p>
"""
        try:
            audit_service.log(user, interface='Chat', action='Read-only Intern Summary', target_type='Intern', target_name=intern_name, status='Success')
        except Exception:
            pass
        return {'ok': True, 'intern': intern_name, 'sheet': sheet_name, 'workbook': str(wb_path), 'html': html, 'candidates': candidates}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
