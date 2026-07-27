from pathlib import Path

root = Path(__file__).resolve().parent
web_app = root / 'web_app.py'
web_dir = root / 'web'
eval_dir = root / 'tracker_evaluation'
readme = root / 'README.md'

if not web_app.exists():
    raise SystemExit('web_app.py not found. Run this patch inside intern_tracker_system_v0.')
web_dir.mkdir(exist_ok=True)
eval_dir.mkdir(exist_ok=True)

# -----------------------------------------------------------------------------
# 1) tracker_evaluation package
# -----------------------------------------------------------------------------
(eval_dir / '__init__.py').write_text('', encoding='utf-8')

(eval_dir / 'evaluation_service.py').write_text(r"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
import json
import uuid

from openpyxl import load_workbook

try:
    from tracker_excel.renderer.parser import parse_workbook
except Exception:
    parse_workbook = None

try:
    from tracker_config.settings import load_settings
    from tracker_llm.providers import build_provider
except Exception:
    load_settings = None
    build_provider = None

BASE_DIR = Path(__file__).resolve().parents[1]
UPLOAD_DIR = BASE_DIR / 'uploads'
OUTPUT_DIR = BASE_DIR / 'outputs'
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

SUBJECTIVE_CRITERIA = [
    'Skills Acquired (plan themes)',
    'Final / Main Project',
    'Real-World Scenario Support',
    'Meets Deadlines',
    'Problem Solving',
    'Creative Thinking',
    'Out-of-the-Box Thinking',
    'Thinking Structure (reasoning)',
    'Decision Making',
    'Communication',
    'Group Participation / Teamwork',
    'Compliance (security/process/reporting)',
    'Ownership & Initiative',
    'Learning Agility / Self-Learning',
    'Documentation Quality',
    'Reliability & Consistency',
]

RUBRICS = {
    'Skills Acquired (plan themes)': '0=not started, 3=works with help, 5=independent/expert across plan skill areas.',
    'Final / Main Project': '0=not attempted, 2=partial/heavy help, 3=working/meets scope, 4=solid/minor help, 5=excellent demo-ready.',
    'Real-World Scenario Support': '0=no contribution, 2=observed only, 3=handled with guidance, 4=mostly independent, 5=led/resolved with strong RCA.',
    'Meets Deadlines': '0=chronically late, 2=often slips, 3=mostly on time, 4=reliably on time, 5=consistently early/on time.',
    'Problem Solving': '0=cannot progress alone, 2=needs constant help, 3=solves routine issues, 4=solves most alone, 5=diagnoses complex problems independently.',
    'Creative Thinking': '0=none, 2=rarely, 3=some original ideas, 4=frequently, 5=consistently inventive solutions.',
    'Out-of-the-Box Thinking': '0=strictly literal, 2=seldom, 3=occasional novel angle, 4=often reframes, 5=non-obvious approaches.',
    'Thinking Structure (reasoning)': '0=disorganised, 2=scattered, 3=logical with prompting, 4=clear steps, 5=structured first-principles reasoning.',
    'Decision Making': '0=avoids/guesses, 2=hesitant, 3=reasonable with data, 4=weighs options, 5=sound judgment on trade-offs.',
    'Communication': '0=unclear/absent, 2=minimal, 3=clear routine updates, 4=proactive, 5=articulate and audience-aware.',
    'Group Participation / Teamwork': '0=disengaged, 2=passive, 3=cooperative, 4=active contributor, 5=elevates team/helps peers.',
    'Compliance (security/process/reporting)': '0=ignores policy, 2=frequent lapses, 3=follows with reminders, 4=disciplined, 5=audit-ready/reports weekly.',
    'Ownership & Initiative': '0=waits to be told, 2=minimal, 3=owns assigned work, 4=anticipates, 5=proactively takes on more.',
    'Learning Agility / Self-Learning': '0=struggles with new topics, 2=slow, 3=learns with support, 4=self-directed, 5=rapid independent learner.',
    'Documentation Quality': '0=none, 2=sparse, 3=usable notes, 4=clear, 5=reusable docs/runbooks.',
    'Reliability & Consistency': '0=unreliable, 2=uneven, 3=steady, 4=dependable, 5=consistent and dependable throughout.',
}

QUESTIONS = {
    'Skills Acquired (plan themes)': 'Across the plan skill areas, what proficiency did the intern demonstrate? Mention independent areas and areas needing help.',
    'Final / Main Project': 'How complete and demo-ready was the final/main project, and how much guidance was required?',
    'Real-World Scenario Support': 'How did the intern contribute to the real-world scenario? Did the intern observe, assist, handle with guidance, or lead?',
    'Meets Deadlines': 'Did the intern complete assigned daily work and weekly projects on time? Mention delays or consistency.',
    'Problem Solving': 'When blocked, how independently did the intern diagnose and solve issues?',
    'Creative Thinking': 'Did the intern suggest original ideas, improvements, or alternative implementations?',
    'Out-of-the-Box Thinking': 'Did the intern reframe problems or suggest non-obvious approaches?',
    'Thinking Structure (reasoning)': 'How structured was the intern’s reasoning and explanation of steps?',
    'Decision Making': 'How well did the intern make trade-off decisions using available data?',
    'Communication': 'How clear, timely, and proactive were the intern’s updates?',
    'Group Participation / Teamwork': 'How well did the intern participate with the team and help peers?',
    'Compliance (security/process/reporting)': 'How consistently did the intern follow process, security rules, reporting, and documentation expectations?',
    'Ownership & Initiative': 'Did the intern own assigned work and proactively take next steps?',
    'Learning Agility / Self-Learning': 'How quickly did the intern learn new topics and self-study?',
    'Documentation Quality': 'How useful and reusable were the intern’s notes, reports, and documentation?',
    'Reliability & Consistency': 'How dependable and consistent was the intern throughout the evaluation period?',
}


def save_upload(upload_file, prefix: str) -> Path:
    suffix = Path(upload_file.filename or 'upload.xlsx').suffix or '.xlsx'
    safe_name = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{suffix}"
    path = UPLOAD_DIR / safe_name
    with path.open('wb') as f:
        f.write(upload_file.file.read())
    return path


def normalize_name(name: str) -> str:
    return ''.join(ch.lower() for ch in (name or '') if ch.isalnum() or ch.isspace()).strip()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize_name(a), normalize_name(b)).ratio()


def get_tracker_interns(tracker_path: str) -> list[dict[str, Any]]:
    if parse_workbook is None:
        return []
    data = parse_workbook(tracker_path)
    interns = []
    for item in getattr(data, 'interns', []) or []:
        start = item.main_row[3] if len(item.main_row) > 3 else ''
        end = item.main_row[4] if len(item.main_row) > 4 else ''
        interns.append({
            'name': item.name,
            'sheet': getattr(item, 'sheet_name', item.name),
            'plan': getattr(item, 'plan_name', '') or '',
            'start': start.strftime('%Y-%m-%d') if hasattr(start, 'strftime') else str(start or ''),
            'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else str(end or ''),
        })
    return interns


def get_eval_scorecards(eval_path: str) -> list[dict[str, Any]]:
    wb = load_workbook(eval_path, data_only=False)
    cards = []
    for ws in wb.worksheets:
        if ws.title.startswith('SC - '):
            cards.append({'name': ws.title.replace('SC - ', '').strip(), 'sheet': ws.title})
    return cards


def match_candidates(tracker_intern: str, eval_cards: list[dict[str, Any]], limit: int = 5) -> list[dict[str, Any]]:
    scored = []
    for card in eval_cards:
        score = similarity(tracker_intern, card['name'])
        scored.append({**card, 'match': round(score * 100, 1)})
    return sorted(scored, key=lambda x: x['match'], reverse=True)[:limit]


def status_done(value: Any) -> bool:
    return str(value or '').strip().lower() in {'completed', 'complete', 'done', 'closed'}


def band_score(completed: int, planned: int) -> int:
    if planned <= 0 or completed <= 0:
        return 0
    pct = completed / planned
    if pct < 0.20:
        return 1
    if pct < 0.50:
        return 2
    if pct < 0.70:
        return 3
    if pct < 0.90:
        return 4
    return 5


def get_tracker_metrics(tracker_path: str, intern_name: str) -> dict[str, Any]:
    if parse_workbook is None:
        return {}
    data = parse_workbook(tracker_path)
    intern = None
    for item in getattr(data, 'interns', []) or []:
        if normalize_name(item.name) == normalize_name(intern_name):
            intern = item
            break
    if not intern:
        # closest tracker intern fallback
        best = None
        best_score = 0
        for item in getattr(data, 'interns', []) or []:
            sc = similarity(item.name, intern_name)
            if sc > best_score:
                best = item; best_score = sc
        intern = best
    if not intern:
        return {}

    tasks = getattr(intern, 'tasks', []) or []
    projects = getattr(intern, 'projects', []) or []
    daily_planned = len(tasks)
    daily_done = sum(1 for row in tasks if len(row) > 4 and status_done(row[4]))
    weekly_planned = len(projects)
    weekly_done = sum(1 for row in projects if len(row) > 5 and status_done(row[5]))

    start = intern.main_row[3] if len(intern.main_row) > 3 else ''
    end = intern.main_row[4] if len(intern.main_row) > 4 else ''
    plan = getattr(intern, 'plan_name', '') or ''
    main_project = intern.main_row[0] if len(intern.main_row) > 0 else ''
    scenario = intern.scenario_row[0] if len(intern.scenario_row) > 0 else ''

    return {
        'matched_tracker_name': intern.name,
        'plan': plan,
        'start': start.strftime('%Y-%m-%d') if hasattr(start, 'strftime') else str(start or ''),
        'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else str(end or ''),
        'daily_done': daily_done,
        'daily_planned': daily_planned,
        'daily_pct': daily_done / daily_planned if daily_planned else 0,
        'weekly_done': weekly_done,
        'weekly_planned': weekly_planned,
        'weekly_pct': weekly_done / weekly_planned if weekly_planned else 0,
        'daily_score': band_score(daily_done, daily_planned),
        'weekly_score': band_score(weekly_done, weekly_planned),
        'main_project': str(main_project or ''),
        'scenario': str(scenario or ''),
    }


def build_questions() -> list[dict[str, Any]]:
    return [
        {
            'criterion': c,
            'question': QUESTIONS[c],
            'rubric': RUBRICS[c],
            'weight_note': 'Score 0–5. Admin can edit final score before writing.'
        }
        for c in SUBJECTIVE_CRITERIA
    ]


def _llm_provider():
    if not load_settings or not build_provider:
        return None
    try:
        settings = load_settings('.env')
        if settings.ai_provider.lower() == 'mock':
            return None
        return build_provider(settings)
    except Exception:
        return None


def heuristic_score(answer: str) -> int:
    a = (answer or '').lower()
    if any(w in a for w in ['outstanding', 'excellent', 'independent', 'led', 'demo-ready', 'demo ready', 'expert']):
        return 5
    if any(w in a for w in ['exceeds', 'mostly independent', 'proactive', 'solid', 'minor help', 'reliable']):
        return 4
    if any(w in a for w in ['meets', 'with guidance', 'routine', 'average', 'mostly on time', 'working']):
        return 3
    if any(w in a for w in ['partial', 'needs help', 'needs constant', 'often', 'below', 'observed']):
        return 2
    if any(w in a for w in ['poor', 'rarely', 'minimal', 'late', 'struggled']):
        return 1
    return 3 if answer.strip() else 0


def suggest_score(criterion: str, answer: str, tracker_context: dict[str, Any]) -> dict[str, Any]:
    provider = _llm_provider()
    rubric = RUBRICS.get(criterion, '')
    if provider:
        prompt = f'''
You are assisting an admin with an intern evaluation. Return ONLY JSON.
Criterion: {criterion}
Rubric: {rubric}
Tracker context: {json.dumps(tracker_context, default=str)}
Evaluator answer: {answer}

Return shape:
{{"score": 0-5, "rationale": "brief specific rationale based only on the answer and tracker context"}}
'''
        try:
            data = provider.complete_json(prompt)
            score = int(data.get('score', heuristic_score(answer)))
            score = max(0, min(5, score))
            return {'score': score, 'rationale': str(data.get('rationale', '')).strip() or 'Suggested from evaluator answer.'}
        except Exception:
            pass
    score = heuristic_score(answer)
    return {'score': score, 'rationale': 'Suggested from evaluator answer using rubric keywords. Admin should review/edit.'}


def _find_cell(ws, text: str):
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or '').strip() == text:
                return cell
    return None


def _write_below_label(ws, label: str, value: Any):
    cell = _find_cell(ws, label)
    if cell:
        ws.cell(row=cell.row + 1, column=cell.column).value = value


def _write_after_label(ws, label: str, value: Any, offset: int = 1):
    cell = _find_cell(ws, label)
    if cell:
        ws.cell(row=cell.row, column=cell.column + offset).value = value


def _write_criterion(ws, criterion: str, score: Any, comment: str = ''):
    cell = _find_cell(ws, criterion)
    if not cell:
        return
    # Expected layout: Criterion in column B, Score in D, Evidence/comments in G.
    ws.cell(row=cell.row, column=cell.column + 2).value = int(score)
    if comment:
        ws.cell(row=cell.row, column=cell.column + 5).value = comment


def finalize_evaluation(eval_path: str, eval_sheet: str, tracker_metrics: dict[str, Any], scores: dict[str, Any], comments: dict[str, str], strengths: str, development: str, remark: str) -> Path:
    wb = load_workbook(eval_path)
    if eval_sheet not in wb.sheetnames:
        raise ValueError(f'Evaluation sheet not found: {eval_sheet}')
    ws = wb[eval_sheet]

    # Delivery snapshot fields already exist in workbook. Do not add fields.
    _write_below_label(ws, 'Daily tasks completed', tracker_metrics.get('daily_done', 0))
    _write_below_label(ws, 'Daily tasks planned', tracker_metrics.get('daily_planned', 0))
    _write_below_label(ws, 'Daily completion %', tracker_metrics.get('daily_pct', 0))
    _write_below_label(ws, 'Weekly projects done', tracker_metrics.get('weekly_done', 0))
    _write_below_label(ws, 'Weekly projects planned', tracker_metrics.get('weekly_planned', 0))

    # Auto scores for delivery completion criteria.
    _write_criterion(ws, 'Daily Tasks Completion', tracker_metrics.get('daily_score', 0), 'Auto-filled from tracker daily task completion.')
    _write_criterion(ws, 'Weekly Tasks / Projects', tracker_metrics.get('weekly_score', 0), 'Auto-filled from tracker weekly project completion.')

    # Admin-approved scores from interview questions.
    for c, score in scores.items():
        _write_criterion(ws, c, score, comments.get(c, ''))

    # Existing text areas.
    _write_below_label(ws, 'Key strengths', strengths)
    _write_below_label(ws, 'Development areas', development)
    _write_below_label(ws, 'Final manager remark', remark)

    out = OUTPUT_DIR / f"Evaluated_{Path(eval_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    return out
""", encoding='utf-8')

# -----------------------------------------------------------------------------
# 2) Evaluation UI page
# -----------------------------------------------------------------------------
(web_dir / 'evaluation.html').write_text(r'''<!doctype html>
<html>
<head>
  <title>Intern Evaluation</title>
  <style>
    body{font-family:Arial,sans-serif;background:#f4f6fb;margin:0;color:#1f2937}header{background:#305496;color:white;padding:18px 28px;display:flex;justify-content:space-between;align-items:center}header a{color:white;font-weight:700;margin-left:14px}main{max-width:1280px;margin:0 auto;padding:20px}.card{background:white;border:1px solid #d9e2ef;border-radius:16px;padding:18px;box-shadow:0 4px 16px rgba(15,23,42,.06);margin-bottom:16px}.grid{display:grid;grid-template-columns:repeat(2,minmax(260px,1fr));gap:16px}.steps{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px}.step{padding:7px 12px;border-radius:999px;background:#e2e8f0;color:#334155;font-size:13px;font-weight:700}.step.active{background:#305496;color:white}input,textarea,select{padding:10px;border:1px solid #d9e2ef;border-radius:9px;font:inherit;width:100%;box-sizing:border-box}textarea{min-height:90px}label{display:flex;flex-direction:column;gap:6px;font-weight:700;margin:8px 0}button{background:#305496;color:white;border:none;border-radius:10px;padding:10px 14px;font-weight:700;cursor:pointer;margin:3px}.success{background:#166534}.danger{background:#991b1b}.muted{color:#64748b;font-size:13px}.match-card{border:1px solid #dbeafe;background:#f8fbff;border-radius:14px;padding:14px;margin:10px 0}.score-row{display:grid;grid-template-columns:1.3fr 90px 1.4fr;gap:10px;align-items:start;border-bottom:1px solid #e5e7eb;padding:10px 0}.pill{display:inline-block;padding:4px 9px;border-radius:999px;background:#dbeafe;color:#1d4ed8;font-weight:700;font-size:12px}.hidden{display:none}.summary{display:grid;grid-template-columns:repeat(5,1fr);gap:10px}.metric{background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;padding:12px}.metric b{display:block;font-size:20px;color:#1d4ed8}
  </style>
</head>
<body>
<header><h2>Intern Evaluation</h2><div><a href="/chat">Chat</a><a href="/users">Users</a><a href="/logs">Logs</a><a href="/profile">Profile</a><a href="/logout">Logout</a></div></header>
<main>
  <div class="steps"><span class="step active" id="s1">1 Upload</span><span class="step" id="s2">2 Match Intern</span><span class="step" id="s3">3 Questions</span><span class="step" id="s4">4 Review</span><span class="step" id="s5">5 Download</span></div>

  <section id="upload" class="card">
    <h3>Upload workbooks</h3>
    <p class="muted">Upload the tracker workbook generated by the system and the evaluation framework workbook. No fields are added to the framework workbook.</p>
    <div class="grid">
      <label>Tracker workbook<input id="trackerFile" type="file" accept=".xlsx"></label>
      <label>Evaluation framework workbook<input id="evalFile" type="file" accept=".xlsx"></label>
    </div>
    <button onclick="startUpload()">Upload & Detect Interns</button>
    <p id="uploadMsg" class="muted"></p>
  </section>

  <section id="match" class="card hidden">
    <h3>Select tracker intern and evaluation scorecard</h3>
    <div class="grid"><label>Tracker intern<select id="internSelect" onchange="showMatches()"></select></label><div><h4>Closest scorecard matches</h4><div id="matchCards"></div></div></div>
    <button onclick="loadQuestions()">Continue to Questions</button>
  </section>

  <section id="questions" class="card hidden">
    <h3 id="qTitle">Guided evaluation</h3>
    <div class="summary" id="metrics"></div>
    <div id="questionBox"></div>
  </section>

  <section id="review" class="card hidden">
    <h3>Review scores before writing workbook</h3>
    <p class="muted">Admin can edit final scores and rationale. The workbook is not updated until Finalize Evaluation is clicked.</p>
    <div id="reviewRows"></div>
    <label>Key strengths<textarea id="strengths"></textarea></label>
    <label>Development areas<textarea id="development"></textarea></label>
    <label>Final manager remark<textarea id="remark"></textarea></label>
    <button class="success" onclick="finalizeEval()">Finalize Evaluation</button>
  </section>

  <section id="download" class="card hidden">
    <h3>Evaluation complete</h3>
    <p id="doneMsg"></p>
    <a id="downloadLink"><button class="success">Download Evaluated Workbook</button></a>
  </section>
</main>
<script>
let sessionId='', trackerInterns=[], evalCards=[], selectedEvalSheet='', questions=[], qIndex=0, answers={}, finalScores={}, rationales={}, trackerMetrics={};
function step(n){[1,2,3,4,5].forEach(i=>document.getElementById('s'+i).classList.toggle('active',i===n));}
function show(id){['upload','match','questions','review','download'].forEach(x=>document.getElementById(x).classList.add('hidden'));document.getElementById(id).classList.remove('hidden');}
async function startUpload(){const fd=new FormData();fd.append('tracker',trackerFile.files[0]);fd.append('evaluation',evalFile.files[0]);uploadMsg.textContent='Uploading...';const r=await fetch('/api/evaluation/upload',{method:'POST',body:fd});const d=await r.json();if(!d.ok){uploadMsg.textContent=d.error||'Upload failed';return;}sessionId=d.session_id;trackerInterns=d.tracker_interns;evalCards=d.eval_cards;internSelect.innerHTML=trackerInterns.map(x=>`<option value="${x.name}">${x.name}</option>`).join('');show('match');step(2);showMatches();}
function showMatches(){const name=internSelect.value;const scored=evalCards.map(c=>({...c,match:similarity(name,c.name)})).sort((a,b)=>b.match-a.match).slice(0,5);matchCards.innerHTML=scored.map((m,i)=>`<div class="match-card"><b>${m.name}</b><br><span class="muted">Sheet: ${m.sheet}</span><br><span class="pill">${Math.round(m.match*100)}% match</span><br><button onclick="chooseSheet('${m.sheet.replaceAll("'","\\'")}')">Use this scorecard</button></div>`).join('');if(scored[0])selectedEvalSheet=scored[0].sheet;}
function chooseSheet(sheet){selectedEvalSheet=sheet;document.querySelectorAll('.match-card').forEach(c=>c.style.outline='');event.target.closest('.match-card').style.outline='3px solid #305496';}
function similarity(a,b){function n(x){return String(x).toLowerCase().replace(/[^a-z0-9 ]/g,'').trim()}a=n(a);b=n(b);if(a===b)return 1;let same=0;for(const ch of a){if(b.includes(ch))same++;}return Math.max(same/Math.max(a.length,1), same/Math.max(b.length,1))*0.7;}
async function loadQuestions(){const r=await fetch('/api/evaluation/questions',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({session_id:sessionId,intern_name:internSelect.value,eval_sheet:selectedEvalSheet})});const d=await r.json();if(!d.ok){alert(d.error);return;}questions=d.questions;trackerMetrics=d.metrics;qIndex=0;renderMetrics();show('questions');step(3);renderQuestion();}
function renderMetrics(){metrics.innerHTML=[['Daily done',trackerMetrics.daily_done+'/'+trackerMetrics.daily_planned],['Daily %',Math.round((trackerMetrics.daily_pct||0)*100)+'%'],['Weekly done',trackerMetrics.weekly_done+'/'+trackerMetrics.weekly_planned],['Daily score',trackerMetrics.daily_score],['Weekly score',trackerMetrics.weekly_score]].map(x=>`<div class="metric"><span>${x[0]}</span><b>${x[1]}</b></div>`).join('');}
function renderQuestion(){const q=questions[qIndex];if(!q){renderReview();return;}qTitle.textContent=`Criterion ${qIndex+1} of ${questions.length}: ${q.criterion}`;questionBox.innerHTML=`<div class="card"><p><b>Rubric:</b> ${q.rubric}</p><p><b>Question:</b> ${q.question}</p><label>Evaluator answer<textarea id="ans">${answers[q.criterion]||''}</textarea></label><button onclick="suggestScore()">Suggest Score</button><button onclick="skipQuestion()">Skip</button><div id="suggestion"></div></div>`;}
async function suggestScore(){const q=questions[qIndex];answers[q.criterion]=ans.value;const r=await fetch('/api/evaluation/suggest',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({criterion:q.criterion,answer:ans.value,metrics:trackerMetrics})});const d=await r.json();if(!d.ok){alert(d.error);return;}finalScores[q.criterion]=d.score;rationales[q.criterion]=d.rationale;suggestion.innerHTML=`<div class="match-card"><b>Suggested score:</b> ${d.score}/5<br><b>Rationale:</b> ${d.rationale}<br><button class="success" onclick="nextQuestion()">Accept & Next</button><button onclick="editSuggestion()">Edit Score</button></div>`;}
function editSuggestion(){const q=questions[qIndex];suggestion.innerHTML+=`<label>Final score<input id="manualScore" type="number" min="0" max="5" value="${finalScores[q.criterion]||0}"></label><label>Rationale<textarea id="manualRat">${rationales[q.criterion]||''}</textarea></label><button class="success" onclick="saveManual()">Save & Next</button>`;}
function saveManual(){const q=questions[qIndex];finalScores[q.criterion]=Number(manualScore.value||0);rationales[q.criterion]=manualRat.value;nextQuestion();}
function skipQuestion(){const q=questions[qIndex];answers[q.criterion]=ans.value;finalScores[q.criterion]=finalScores[q.criterion] ?? 0;rationales[q.criterion]=rationales[q.criterion] || 'Skipped or not enough evidence.';nextQuestion();}
function nextQuestion(){qIndex++;renderQuestion();}
function renderReview(){show('review');step(4);reviewRows.innerHTML=questions.map(q=>`<div class="score-row"><div><b>${q.criterion}</b><p class="muted">${q.rubric}</p></div><input type="number" min="0" max="5" value="${finalScores[q.criterion]??0}" data-score="${q.criterion}"><textarea data-rat="${q.criterion}">${rationales[q.criterion]||''}</textarea></div>`).join('');}
async function finalizeEval(){document.querySelectorAll('[data-score]').forEach(i=>finalScores[i.dataset.score]=Number(i.value||0));document.querySelectorAll('[data-rat]').forEach(t=>rationales[t.dataset.rat]=t.value);const r=await fetch('/api/evaluation/finalize',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({session_id:sessionId,intern_name:internSelect.value,eval_sheet:selectedEvalSheet,metrics:trackerMetrics,scores:finalScores,comments:rationales,strengths:strengths.value,development:development.value,remark:remark.value})});const d=await r.json();if(!d.ok){alert(d.error);return;}show('download');step(5);doneMsg.textContent='Evaluation workbook generated.';downloadLink.href=d.download;}
</script>
</body>
</html>''', encoding='utf-8')

# -----------------------------------------------------------------------------
# 3) web_app routes
# -----------------------------------------------------------------------------
s = web_app.read_text(encoding='utf-8')

# Ensure imports are present after future import line, not before.
if 'from tracker_evaluation.evaluation_service import' not in s:
    import_line = "from tracker_evaluation.evaluation_service import save_upload, get_tracker_interns, get_eval_scorecards, match_candidates, get_tracker_metrics, build_questions, suggest_score, finalize_evaluation\n"
    lines = s.splitlines()
    insert_at = 1 if lines and lines[0].strip() == 'from __future__ import annotations' else 0
    lines.insert(insert_at, import_line.rstrip())
    s = '\n'.join(lines) + '\n'

# Ensure File/UploadFile imports.
if 'UploadFile' not in s or 'File' not in s:
    s = s.replace('from fastapi import ', 'from fastapi import UploadFile, File, ', 1) if 'from fastapi import ' in s else 'from fastapi import UploadFile, File\n' + s

if 'EVAL_SESSIONS = {}' not in s:
    s += "\n\nEVAL_SESSIONS = {}\n"

if "@app.get('/evaluation'" not in s:
    routes = r'''

@app.get('/evaluation', response_class=HTMLResponse)
def evaluation_page(request: Request):
    user = current_user_from_request(request)
    if not user:
        return RedirectResponse('/login')
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return RedirectResponse('/chat')
    return (BASE_DIR / 'web' / 'evaluation.html').read_text(encoding='utf-8')

@app.post('/api/evaluation/upload')
def api_evaluation_upload(request: Request, tracker: UploadFile = File(...), evaluation: UploadFile = File(...)):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    try:
        tracker_path = save_upload(tracker, 'eval_tracker')
        eval_path = save_upload(evaluation, 'eval_framework')
        tracker_interns = get_tracker_interns(str(tracker_path))
        eval_cards = get_eval_scorecards(str(eval_path))
        session_id = __import__('uuid').uuid4().hex
        EVAL_SESSIONS[session_id] = {'tracker': str(tracker_path), 'evaluation': str(eval_path), 'user': user.get('email')}
        audit_service.log(user, interface='Evaluation', action='Upload Evaluation Files', status='Success', summary=f'{len(tracker_interns)} tracker interns, {len(eval_cards)} scorecards')
        return {'ok': True, 'session_id': session_id, 'tracker_interns': tracker_interns, 'eval_cards': eval_cards}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})

@app.post('/api/evaluation/questions')
def api_evaluation_questions(request: Request, payload: dict):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    sess = EVAL_SESSIONS.get(payload.get('session_id'))
    if not sess:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Evaluation session not found'})
    metrics = get_tracker_metrics(sess['tracker'], payload.get('intern_name',''))
    questions = build_questions()
    return {'ok': True, 'metrics': metrics, 'questions': questions}

@app.post('/api/evaluation/suggest')
def api_evaluation_suggest(request: Request, payload: dict):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    try:
        result = suggest_score(payload.get('criterion',''), payload.get('answer',''), payload.get('metrics') or {})
        return {'ok': True, **result}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})

@app.post('/api/evaluation/finalize')
def api_evaluation_finalize(request: Request, payload: dict):
    user = require_login(request)
    if user.get('role') not in {'Super Admin', 'Admin'}:
        return JSONResponse(status_code=403, content={'ok': False, 'error': 'Admin or Super Admin only'})
    sess = EVAL_SESSIONS.get(payload.get('session_id'))
    if not sess:
        return JSONResponse(status_code=404, content={'ok': False, 'error': 'Evaluation session not found'})
    try:
        out = finalize_evaluation(sess['evaluation'], payload.get('eval_sheet'), payload.get('metrics') or {}, payload.get('scores') or {}, payload.get('comments') or {}, payload.get('strengths',''), payload.get('development',''), payload.get('remark',''))
        audit_service.log(user, interface='Evaluation', action='Finalize Evaluation', target_type='Intern', target_name=payload.get('intern_name',''), output_workbook=str(out), status='Success')
        return {'ok': True, 'output_path': str(out), 'download': '/download?path=' + str(out)}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
'''
    s += routes

web_app.write_text(s, encoding='utf-8')

# -----------------------------------------------------------------------------
# 4) Add Evaluation nav link for Admin/Super Admin via JS on key pages
# -----------------------------------------------------------------------------
nav_js = r'''
<script id="v70-evaluation-nav">
async function v70EvaluationNav(){try{const r=await fetch('/api/me');const d=await r.json();const role=((d.user&&d.user.role)||'').toLowerCase();if(role==='admin'||role==='super admin'){let nav=document.querySelector('header .nav')||document.querySelector('header div:last-child')||document.querySelector('header');if(nav&&!document.querySelector('header a[href="/evaluation"]')){const a=document.createElement('a');a.href='/evaluation';a.textContent='Evaluation';a.style.color='white';a.style.fontWeight='700';a.style.marginLeft='14px';nav.appendChild(a);}}}catch(e){}}
v70EvaluationNav();
</script>
'''
for page in ['chat.html','users.html','logs.html','tasks.html','profile.html']:
    p = web_dir / page
    if p.exists():
        html = p.read_text(encoding='utf-8')
        if 'v70-evaluation-nav' not in html:
            html = html.replace('</body>', nav_js + '\n</body>')
            p.write_text(html, encoding='utf-8')

# Compile checks
try:
    import py_compile
    py_compile.compile(str(eval_dir / 'evaluation_service.py'), doraise=True)
    py_compile.compile(str(web_app), doraise=True)
except Exception as e:
    raise SystemExit(f'Compile check failed: {e}')

if readme.exists():
    readme.write_text(readme.read_text(encoding='utf-8') + r'''

## v0.70 Evaluation workflow

- Added Admin/Super Admin-only `/evaluation` page.
- Admin uploads tracker workbook and evaluation framework workbook.
- System detects tracker interns and evaluation scorecards.
- Closest matching scorecards are shown as professional cards.
- System auto-fills delivery metrics from tracker.
- LLM/evaluation assistant asks subjective criterion questions and proposes 0–5 scores with rationale.
- Admin can edit scores/rationale before finalizing.
- Existing evaluation workbook structure is used; no new fields are added.
- Finalized scorecard is written to a downloadable evaluated workbook.
''', encoding='utf-8')

print('v0.70 evaluation workflow patch applied successfully.')
