"""
Patch v0.92 - Chat recovery + direct read-only intern summary

Apply from intern_tracker_system_v0 root:
    python patch_v92_chat_recovery_readonly_summary.py

Why:
- v87-v91 UI patches disrupted the approval panel / side panel behavior.
- Read-only prompts like "how is Bilal doing?" still enter proposal approval flow.

What this patch does:
1. Removes earlier experimental chat UI/auto-approval/debug injected scripts/styles from chat.html.
2. Restores the approval area so workbook-changing actions can display approval controls normally.
3. Adds a dedicated read-only summary API endpoint:
       POST /api/readonly/intern-summary
4. Adds a frontend interceptor for read-only questions like:
       how is Bilal doing?
       how is Bilal Ahmad Khan doing?
       show Bilal progress
   These are answered directly and do not enter proposal approval flow.
5. Does not auto-approve workbook-changing commands.
6. Does not hardcode any intern name.
"""
from __future__ import annotations

from pathlib import Path
import re
import textwrap

ROOT = Path(__file__).resolve().parent
WEB_APP = ROOT / "web_app.py"
CHAT_HTML_CANDIDATES = [ROOT / "web" / "chat.html", ROOT / "chat.html", ROOT / "templates" / "chat.html"]
README = ROOT / "README.md"

if not WEB_APP.exists():
    raise SystemExit("web_app.py not found. Run from intern_tracker_system_v0 root folder.")
chat_html = next((p for p in CHAT_HTML_CANDIDATES if p.exists()), None)
if chat_html is None:
    raise SystemExit("chat.html not found. Run from intern_tracker_system_v0 root folder.")

# -----------------------------------------------------------------------------
# 1) Clean experimental chat UI scripts/styles that caused layout/proposal issues
# -----------------------------------------------------------------------------
h = chat_html.read_text(encoding="utf-8")
remove_ids = [
    "v84-readonly-summary-no-proposal",
    "v85-summary-no-approval-frontend",
    "v86-auto-execute-readonly-summary-ui",
    "v87-chat-ui-compact-style",
    "v87-chat-ui-compact-script",
    "v88-chat-empty-approval-collapse-style",
    "v88-chat-empty-approval-collapse-script",
    "v89-chat-review-precise-collapse-style",
    "v89-chat-review-precise-collapse-script",
    "v90-chat-layout-debug-style",
    "v90-chat-layout-debug-script",
    "v91-chat-side-panel-block-style",
    "v91-chat-side-panel-block-script",
]
for rid in remove_ids:
    h = re.sub(rf"\n?<(script|style)[^>]+id=[\"']{re.escape(rid)}[\"'][\s\S]*?</\1>\s*", "\n", h, flags=re.I)

# Remove old problematic CSS classes if they were static in class attributes.
h = h.replace(" v87-approval-active-card", "")
h = h.replace(" v87-current-workbook-card", "")
h = h.replace(" v91-empty-approval-side", "")
h = h.replace(" v91-active-approval-side", "")

# -----------------------------------------------------------------------------
# 2) Add frontend read-only summary interceptor
# -----------------------------------------------------------------------------
frontend_block = r'''

<!-- v0.92 direct read-only intern summary interceptor -->
<style id="v92-chat-recovery-style">
  /* Restore the side panel to normal behavior after experimental patches. */
  aside.panel.side,
  .panel.side {
    display: flex;
    height: auto;
    min-height: 0;
    max-height: none;
    overflow: visible;
  }

  header a,
  .topbar a,
  .navbar a,
  .app-header a {
    margin-left: 18px !important;
    display: inline-block !important;
    white-space: nowrap !important;
  }

  .v92-summary-bubble {
    max-width: 760px;
    background: #ffffff;
    border: 1px solid #d9e2ef;
    border-radius: 14px;
    padding: 14px 16px;
    margin: 12px 20px;
    box-shadow: 0 2px 10px rgba(15, 23, 42, 0.05);
    line-height: 1.45;
  }

  .v92-summary-bubble h3 {
    margin: 0 0 8px;
    color: #1f3f75;
  }

  .v92-summary-bubble ul {
    margin-top: 6px;
  }

  .v92-summary-user {
    max-width: 760px;
    background: #dbeafe;
    border-radius: 14px;
    padding: 12px 16px;
    margin: 12px 20px 12px auto;
    line-height: 1.4;
  }
</style>
<script id="v92-readonly-summary-interceptor">
(function(){
  const MUTATION_WORDS = [
    'add intern','extend intern','create plan','make a fresh excel','fresh excel','edit task','update task',
    'add holiday','finalize evaluation','apply plan','create workbook','delete','remove','change status'
  ];

  function norm(v){ return String(v || '').replace(/\s+/g, ' ').trim(); }
  function low(v){ return norm(v).toLowerCase(); }

  function isReadonlySummaryPrompt(text){
    const t = low(text);
    if(!t) return false;
    if(MUTATION_WORDS.some(w => t.includes(w))) return false;
    return /^how\s+is\s+.+\s+doing\??$/.test(t)
        || /^how\s+is\s+.+\??$/.test(t)
        || /^show\s+.+\s+progress\??$/.test(t)
        || /^summar(y|ize|ise)\s+.+\??$/.test(t)
        || /^.+\s+progress\??$/.test(t);
  }

  function findInput(){
    const inputs = Array.from(document.querySelectorAll('textarea, input[type="text"], input:not([type])'));
    return inputs.reverse().find(el => {
      const style = window.getComputedStyle(el);
      return style.display !== 'none' && style.visibility !== 'hidden' && !el.disabled && !el.readOnly;
    });
  }

  function currentWorkbookValue(){
    // Prefer Workbook files select.
    const selects = Array.from(document.querySelectorAll('select'));
    for(const sel of selects){
      const nearby = low((sel.closest('section, aside, div, label') || sel).textContent);
      if(nearby.includes('workbook')){
        const opt = sel.options && sel.selectedIndex >= 0 ? sel.options[sel.selectedIndex] : null;
        return (opt && (opt.value || opt.textContent)) || sel.value || '';
      }
    }
    // Fallback: look for a visible xlsx cell/input.
    const nodes = Array.from(document.querySelectorAll('input, div, span, p'));
    const hit = nodes.find(el => low(el.textContent || el.value).includes('.xlsx'));
    return hit ? norm(hit.value || hit.textContent) : '';
  }

  function conversationContainer(){
    const headings = Array.from(document.querySelectorAll('h1,h2,h3,h4,div')).filter(el => low(el.textContent) === 'conversation');
    if(headings[0]) return headings[0].closest('section, article, div, main') || document.querySelector('main') || document.body;
    return document.querySelector('main') || document.body;
  }

  function appendUser(text){
    const box = document.createElement('div');
    box.className = 'v92-summary-user';
    box.textContent = text;
    conversationContainer().appendChild(box);
  }

  function appendAssistant(html){
    const box = document.createElement('div');
    box.className = 'v92-summary-bubble';
    box.innerHTML = html;
    conversationContainer().appendChild(box);
    box.scrollIntoView({behavior:'smooth', block:'nearest'});
  }

  async function runReadonlySummary(prompt){
    appendUser(prompt);
    appendAssistant('<h3>Progress Summary</h3><p>Generating summary...</p>');
    const bubbles = document.querySelectorAll('.v92-summary-bubble');
    const bubble = bubbles[bubbles.length - 1];
    try{
      const r = await fetch('/api/readonly/intern-summary', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({prompt: prompt, workbook: currentWorkbookValue()})
      });
      const d = await r.json();
      if(d.ok){
        bubble.innerHTML = d.html || ('<h3>Progress Summary</h3><p>' + (d.message || 'Summary generated.') + '</p>');
      }else{
        bubble.innerHTML = '<h3>Progress Summary</h3><p style="color:#991b1b">' + (d.error || 'Could not generate summary.') + '</p>';
      }
    }catch(e){
      bubble.innerHTML = '<h3>Progress Summary</h3><p style="color:#991b1b">Summary failed: ' + e.message + '</p>';
    }
  }

  function interceptSubmit(e){
    const input = findInput();
    if(!input) return;
    const prompt = norm(input.value);
    if(!isReadonlySummaryPrompt(prompt)) return;
    e.preventDefault();
    e.stopPropagation();
    input.value = '';
    input.dispatchEvent(new Event('input', {bubbles:true}));
    runReadonlySummary(prompt);
  }

  document.addEventListener('submit', interceptSubmit, true);
  document.addEventListener('keydown', function(e){
    if(e.key !== 'Enter' || e.shiftKey) return;
    const input = findInput();
    if(!input || document.activeElement !== input) return;
    const prompt = norm(input.value);
    if(!isReadonlySummaryPrompt(prompt)) return;
    e.preventDefault();
    e.stopPropagation();
    input.value = '';
    input.dispatchEvent(new Event('input', {bubbles:true}));
    runReadonlySummary(prompt);
  }, true);
  document.addEventListener('click', function(e){
    const btn = e.target.closest('button, input[type="button"], input[type="submit"]');
    if(!btn) return;
    const label = low(btn.textContent || btn.value);
    if(label !== 'send' && label !== 'submit' && label !== 'ask') return;
    const input = findInput();
    if(!input) return;
    const prompt = norm(input.value);
    if(!isReadonlySummaryPrompt(prompt)) return;
    e.preventDefault();
    e.stopPropagation();
    input.value = '';
    input.dispatchEvent(new Event('input', {bubbles:true}));
    runReadonlySummary(prompt);
  }, true);
})();
</script>
'''

if "v92-readonly-summary-interceptor" not in h:
    h = h.replace("</body>", frontend_block + "\n</body>") if "</body>" in h else h + frontend_block

chat_html.write_text(h, encoding="utf-8")

# -----------------------------------------------------------------------------
# 3) Add backend read-only summary API endpoint
# -----------------------------------------------------------------------------
s = WEB_APP.read_text(encoding="utf-8")

# Ensure openpyxl import availability by importing inside endpoint helpers. Ensure JSONResponse exists.
if "JSONResponse" not in s and "from fastapi.responses" in s:
    s = s.replace("from fastapi.responses import ", "from fastapi.responses import JSONResponse, ", 1)

endpoint = r'''

# v0.92 read-only intern progress summary endpoint
@app.post('/api/readonly/intern-summary')
def api_v92_readonly_intern_summary(request: Request, payload: dict):
    user = require_login(request)
    try:
        from pathlib import Path
        from difflib import SequenceMatcher
        from datetime import datetime, date
        from openpyxl import load_workbook
        import re

        def norm(v):
            return re.sub(r'\s+', ' ', str(v or '').strip())

        def norm_name(v):
            return re.sub(r'[^a-z0-9]+', ' ', str(v or '').lower()).strip()

        def status_done(v):
            return str(v or '').strip().lower() in {'completed', 'complete', 'done'}

        def parse_date(v):
            if hasattr(v, 'date'):
                return v.date()
            return None

        def pct(a, b):
            return (a / b) if b else 0

        def extract_requested_name(prompt):
            p = norm(prompt)
            p2 = re.sub(r'(?i)^how\s+is\s+', '', p)
            p2 = re.sub(r'(?i)\s+doing\??$', '', p2)
            p2 = re.sub(r'(?i)^show\s+', '', p2)
            p2 = re.sub(r'(?i)\s+progress\??$', '', p2)
            p2 = re.sub(r'(?i)^summar(?:y|ize|ise)\s+', '', p2)
            return norm(p2) or p

        def resolve_workbook(value):
            raw = norm(value)
            candidates = []
            if raw:
                cleaned = raw.replace('outputs /', 'outputs/').replace('uploads /', 'uploads/')
                candidates.append(BASE_DIR / cleaned)
                candidates.append(BASE_DIR / 'outputs' / Path(cleaned).name)
                candidates.append(BASE_DIR / 'uploads' / Path(cleaned).name)
                candidates.append(Path(cleaned))
            for c in candidates:
                if c.exists() and c.is_file():
                    return c
            # Fallback to latest output workbook.
            outs = sorted((BASE_DIR / 'outputs').glob('*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True) if (BASE_DIR / 'outputs').exists() else []
            if outs:
                return outs[0]
            ups = sorted((BASE_DIR / 'uploads').glob('*.xlsx'), key=lambda p: p.stat().st_mtime, reverse=True) if (BASE_DIR / 'uploads').exists() else []
            if ups:
                return ups[0]
            return None

        def discover_intern_sheets(wb):
            items = []
            for ws in wb.worksheets:
                title = str(ws.title or '')
                a1 = str(ws['A1'].value or '')
                if 'Intern Tracker' in a1:
                    name = a1.split('—', 1)[-1].split('(', 1)[0].strip() if '—' in a1 else title
                    items.append((name, title))
            return items

        def find_best_intern(wb, requested):
            sheets = discover_intern_sheets(wb)
            req = norm_name(requested)
            if not sheets:
                return None, None, []
            scored = []
            req_tokens = set(req.split())
            for name, sheet in sheets:
                nn = norm_name(name)
                ns = norm_name(sheet)
                tokens = set(nn.split()) | set(ns.split())
                overlap = len(req_tokens & tokens) / max(len(req_tokens), 1)
                ratio = max(SequenceMatcher(None, req, nn).ratio(), SequenceMatcher(None, req, ns).ratio())
                score = max(overlap, ratio)
                scored.append((score, name, sheet))
            scored.sort(reverse=True)
            best = scored[0]
            return best[1], best[2], [{'name': n, 'sheet': sh, 'score': round(sc, 3)} for sc, n, sh in scored[:5]]

        def find_row(ws, label):
            target = str(label).strip().lower()
            for row in ws.iter_rows():
                for cell in row:
                    if str(cell.value or '').strip().lower().startswith(target):
                        return cell.row
            return None

        def parse_daily(ws):
            row = find_row(ws, 'DAILY TASKS')
            if not row:
                return []
            header = row + 1
            headers = {str(ws.cell(header, c).value or '').strip().lower(): c for c in range(1, ws.max_column + 1)}
            c_date = headers.get('date')
            c_week = headers.get('week') or headers.get('week #')
            c_theme = headers.get('theme')
            c_task = headers.get('task description') or headers.get('task')
            c_status = next((c for h, c in headers.items() if 'status' in h), None)
            tasks = []
            if not c_status:
                return tasks
            for r in range(header + 1, ws.max_row + 1):
                first = str(ws.cell(r, 1).value or '').strip().lower()
                if first in {'weekly updates', 'small projects / tasks', 'small projects / tasks  (weekly projects)', 'main project', 'real-world scenario'}:
                    break
                status = ws.cell(r, c_status).value
                if status in (None, ''):
                    continue
                tasks.append({
                    'date': parse_date(ws.cell(r, c_date).value) if c_date else None,
                    'week': ws.cell(r, c_week).value if c_week else '',
                    'theme': ws.cell(r, c_theme).value if c_theme else '',
                    'task': ws.cell(r, c_task).value if c_task else '',
                    'status': status,
                })
            return tasks

        def parse_projects(ws):
            row = None
            for r in range(1, ws.max_row + 1):
                if str(ws.cell(r, 1).value or '').strip().lower().startswith('small projects / tasks'):
                    row = r
                    break
            if not row:
                return []
            header = row + 1
            headers = {str(ws.cell(header, c).value or '').strip().lower(): c for c in range(1, ws.max_column + 1)}
            c_title = headers.get('title')
            c_status = next((c for h, c in headers.items() if 'status' in h), None)
            projects = []
            if not c_status:
                return projects
            for r in range(header + 1, ws.max_row + 1):
                status = ws.cell(r, c_status).value
                if status in (None, ''):
                    continue
                projects.append({'title': ws.cell(r, c_title).value if c_title else '', 'status': status})
            return projects

        prompt = payload.get('prompt', '')
        wb_path = resolve_workbook(payload.get('workbook', ''))
        if not wb_path:
            return JSONResponse(status_code=400, content={'ok': False, 'error': 'No workbook found. Please select or upload a workbook first.'})

        wb = load_workbook(wb_path, data_only=True)
        requested = extract_requested_name(prompt)
        intern_name, sheet_name, candidates = find_best_intern(wb, requested)
        if not sheet_name:
            return JSONResponse(status_code=404, content={'ok': False, 'error': f'No intern tracker sheets found in {wb_path.name}.'})
        ws = wb[sheet_name]
        tasks = parse_daily(ws)
        projects = parse_projects(ws)

        total = len(tasks)
        completed = sum(1 for t in tasks if status_done(t['status']))
        in_progress = sum(1 for t in tasks if str(t['status']).strip().lower() == 'in progress')
        pending = total - completed - in_progress
        completion = pct(completed, total)
        ptotal = len(projects)
        pdone = sum(1 for p in projects if status_done(p['status']))
        pcompletion = pct(pdone, ptotal)

        completed_weeks = []
        pending_weeks = []
        for t in tasks:
            wk = t.get('week')
            if status_done(t['status']):
                completed_weeks.append(wk)
            elif wk not in ('', None):
                pending_weeks.append(wk)
        current_week = pending_weeks[0] if pending_weeks else ('Done' if total and completed == total else '')

        completed_themes = []
        pending_themes = []
        for t in tasks:
            theme = norm(t.get('theme'))
            if not theme:
                continue
            if status_done(t['status']) and theme not in completed_themes:
                completed_themes.append(theme)
            if not status_done(t['status']) and theme not in pending_themes:
                pending_themes.append(theme)

        status_label = 'On Track' if completion >= 0.75 else ('Developing' if completion >= 0.45 else 'Needs Support')
        html = f"""
<h3>{intern_name} - Progress Summary</h3>
<ul>
  <li><b>Workbook:</b> {wb_path.name}</li>
  <li><b>Daily tasks:</b> {completed}/{total} completed ({completion:.0%})</li>
  <li><b>Weekly projects:</b> {pdone}/{ptotal} completed ({pcompletion:.0%})</li>
  <li><b>In progress:</b> {in_progress}</li>
  <li><b>Pending:</b> {pending}</li>
  <li><b>Current/next week:</b> {current_week}</li>
  <li><b>Status:</b> {status_label}</li>
</ul>
<p><b>Completed areas:</b> {', '.join(completed_themes[:4]) or 'Not available from tracker.'}</p>
<p><b>Pending/upcoming areas:</b> {', '.join(pending_themes[:4]) or 'No pending areas found.'}</p>
<p><b>Suggested manager action:</b> Review pending tasks/projects and ask for blockers if progress is below expected pace.</p>
"""
        try:
            audit_service.log(user, interface='Chat', action='Read-only Intern Summary', target_type='Intern', target_name=intern_name, status='Success')
        except Exception:
            pass
        return {'ok': True, 'intern': intern_name, 'sheet': sheet_name, 'workbook': str(wb_path), 'html': html, 'candidates': candidates}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
'''

if "api_v92_readonly_intern_summary" not in s:
    s += endpoint

WEB_APP.write_text(s, encoding="utf-8")

# Compile check
try:
    import py_compile
    py_compile.compile(str(WEB_APP), doraise=True)
except Exception as e:
    raise SystemExit(f"web_app.py compile failed after v0.92: {e}")

if README.exists():
    txt = README.read_text(encoding="utf-8", errors="ignore")
    if "v0.92 Chat recovery + direct read-only intern summary" not in txt:
        README.write_text(txt + textwrap.dedent("""

        ## v0.92 Chat recovery + direct read-only intern summary

        - Removes experimental chat UI/debug/auto-approval patches v84-v91 from chat.html.
        - Restores normal approval panel behavior for workbook-changing commands.
        - Adds direct read-only intern progress summary API and chat interceptor.
        - Prompts like `how is Bilal doing?` are answered directly without approval.
        """), encoding="utf-8")

print("v0.92 chat recovery + direct read-only summary patch applied successfully.")
print(f"Updated: {WEB_APP}")
print(f"Updated: {chat_html}")
