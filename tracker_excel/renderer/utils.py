from datetime import datetime, date
from openpyxl.utils import get_column_letter
from .styles import STYLES, apply_cell, apply_range


def default_visible_page(ws):
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_margins.left = 0.75
    ws.page_margins.right = 0.75
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 1.0


def style_row(ws, row, start_col, end_col, style_name):
    for c in range(start_col, end_col + 1):
        apply_cell(ws.cell(row, c), STYLES[style_name])


def write_row(ws, row, values, style_name=None, start_col=1):
    for idx, value in enumerate(values, start_col):
        cell = ws.cell(row, idx, value)
        if style_name:
            apply_cell(cell, STYLES[style_name])


def set_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def is_date(value):
    return isinstance(value, (datetime, date))


def excel_date_formula(dt):
    if isinstance(dt, datetime):
        return f'DATE({dt.year},{dt.month},{dt.day})'
    if isinstance(dt, date):
        return f'DATE({dt.year},{dt.month},{dt.day})'
    return None


def max_numeric_week(tasks):
    weeks = []
    for t in tasks:
        if len(t) > 1 and isinstance(t[1], int):
            weeks.append(t[1])
    return max(weeks) if weeks else 0


def estimate_wrapped_height(values, widths, min_height=24, base_line_height=15, padding=8, max_height=180):
    """Estimate Excel row height for wrapped text based on text length and column widths."""
    max_lines = 1
    for idx, value in enumerate(values, 1):
        if value is None:
            continue
        text = str(value)
        if not text:
            continue
        # Approx characters per line: Excel width roughly equals chars for Arial 10.
        width = widths.get(idx, 12)
        chars_per_line = max(8, int(width * 0.9))
        lines = 0
        for part in text.splitlines() or [text]:
            lines += max(1, (len(part) + chars_per_line - 1) // chars_per_line)
        max_lines = max(max_lines, lines)
    return min(max(min_height, max_lines * base_line_height + padding), max_height)


def set_dynamic_row_height(ws, row, values, width_by_index, min_height=24, max_height=180):
    ws.row_dimensions[row].height = estimate_wrapped_height(values, width_by_index, min_height=min_height, max_height=max_height)
