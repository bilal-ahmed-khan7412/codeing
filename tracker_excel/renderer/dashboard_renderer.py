from datetime import datetime, date
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font
from .styles import STYLES, apply_cell, GREEN_FILL, GREEN_TEXT, RED_FILL, RED_TEXT, YELLOW_FILL, YELLOW_TEXT, ORANGE_FILL
from .utils import default_visible_page, write_row, style_row, set_widths, excel_date_formula


def _render_empty_dashboard_shell(ws):
    """Render a safe dashboard shell when there are no interns yet."""
    from .utils import write_row, style_row, set_widths
    from .styles import STYLES, apply_cell
    set_widths(ws, {'A':30,'B':16,'C':16,'D':16,'E':16,'F':18,'G':18,'H':18,'I':22,'J':48})
    ws.cell(1,1,'Internship Learning — Dashboard - Systems Limited')
    style_row(ws,1,1,10,'dashboard_title')
    ws.row_dimensions[1].height = 32
    ws.cell(2,1,'No interns added yet. Create/add an intern to populate KPIs and charts.')
    style_row(ws,2,1,10,'subtitle_center')
    ws.cell(4,1,'KPIs — Per Intern')
    style_row(ws,4,1,10,'section')
    write_row(ws,5,['Intern','Total Tasks','Completed','In Progress','Pending','Completion %','On Track?','Current Week','Weekly Projects Done','Days Left'],'table_header')
    write_row(ws,6,['No interns yet','','','','','','','','',''],'body_center')
    ws.cell(9,1,'Manager View — Reporting & Engagement')
    style_row(ws,9,1,7,'section')
    write_row(ws,10,['Intern','Weeks in Plan','Weekly Emails Sent','LM Acknowledged','Reporting %','Learning Health','Attention Needed?'],'table_header')
    write_row(ws,11,['No interns yet','','','','','',''],'body_center')
    ws.cell(14,1,'Overall Status Breakdown')
    style_row(ws,14,1,3,'section')
    write_row(ws,15,['Status','Count'],'table_header')
    write_row(ws,16,['Completed',0],'body_left')
    write_row(ws,17,['In Progress',0],'body_left')
    write_row(ws,18,['Pending',0],'body_left')
    ws.cell(20,1,'Completion % by Intern (data)')
    style_row(ws,20,1,3,'section')
    write_row(ws,21,['Intern','Completion %'],'table_header')
    write_row(ws,22,['No interns yet',0],'body_left')
    apply_cell(ws.cell(22,2), STYLES['pct'])
    ws.cell(25,1,'Tasks Completed per Week (per Intern)')
    style_row(ws,25,1,6,'section')
    write_row(ws,26,['Week'],'table_header')
    for r in range(27,35):
        write_row(ws,r,[f'W{r-26}'],'body_center')
    return ws


def render_dashboard(wb, intern_meta):
    ws = wb.create_sheet('Dashboard', 0)
    default_visible_page(ws)
    if not intern_meta:
        return _render_empty_dashboard_shell(ws)
    set_widths(ws, {'A':30,'B':14,'C':14,'D':14,'E':14,'F':16,'G':16,'H':16,'I':20,'J':48})
    ws.cell(1,1,'Internship Learning — Dashboard - Systems Limited')
    style_row(ws,1,1,10,'dashboard_title')
    ws.row_dimensions[1].height = 31.5
    summary = '  ·  '.join([m['name'] for m in intern_meta])
    ws.cell(2,1,summary)
    style_row(ws,2,1,10,'subtitle_center')

    # KPI section
    ws.cell(4,1,'KPIs — Per Intern')
    style_row(ws,4,1,10,'section')
    headers = ['Intern','Total Tasks','Completed','In Progress','Pending','Completion %','On Track?','Current Week','Weekly Projects Done','Days Left']
    write_row(ws,5,headers,'table_header')
    start = 6
    for idx,m in enumerate(intern_meta, start=start):
        sh = m['sheet']
        ts,te = m['task_start'],m['task_end']
        ps,pe = m['project_start'],m['project_end']
        target_formula = excel_date_formula(m.get('target_end'))
        vals = [m['name'],
                f'=COUNTA(\'{sh}\'!E{ts}:E{te})',
                f'=COUNTIF(\'{sh}\'!E{ts}:E{te},"Completed")',
                f'=COUNTIF(\'{sh}\'!E{ts}:E{te},"In Progress")',
                f'=COUNTIF(\'{sh}\'!E{ts}:E{te},"Pending")',
                f'=IFERROR(C{idx}/B{idx},0)',
                f'=IF(F{idx}>=0.8,"On Track",IF(F{idx}>=0.5,"At Risk","Behind"))',
                f'=IFERROR(INDEX(\'{sh}\'!B{ts}:B{te},MATCH("Pending",\'{sh}\'!E{ts}:E{te},0)),"Done")',
                f'=COUNTIF(\'{sh}\'!F{ps}:F{pe},"Completed")&" / "&COUNTA(\'{sh}\'!F{ps}:F{pe})' if pe >= ps else '0 / 0',
                f'=MAX(0,{target_formula}-TODAY())' if target_formula else '']
        write_row(ws,idx,vals,'body_center')
        apply_cell(ws.cell(idx,1), STYLES['row_label'])
        apply_cell(ws.cell(idx,6), STYLES['pct'])
    total_row = start + len(intern_meta)
    vals = ['TOTAL / AVG', f'=SUM(B{start}:B{total_row-1})', f'=SUM(C{start}:C{total_row-1})', f'=SUM(D{start}:D{total_row-1})', f'=SUM(E{start}:E{total_row-1})', f'=IFERROR(C{total_row}/B{total_row},0)']
    write_row(ws,total_row,vals,'body_center')
    apply_cell(ws.cell(total_row,1), STYLES['row_label'])
    apply_cell(ws.cell(total_row,6), STYLES['pct'])

    # Manager view
    manager_section = total_row + 2
    ws.cell(manager_section,1,'Manager View — Reporting & Engagement')
    style_row(ws,manager_section,1,7,'section')
    m_header = manager_section + 1
    write_row(ws,m_header,['Intern','Weeks in Plan','Weekly Emails Sent','LM Acknowledged','Reporting %','Learning Health','Attention Needed?'],'table_header')
    for n,m in enumerate(intern_meta, start=m_header+1):
        sh=m['sheet']; wsrow = n; kpi_row = start + (n - (m_header+1))
        vals=[m['name'], m['week_count'], f'=COUNTIF(\'{sh}\'!G{m["weekly_start"]}:G{m["weekly_end"]},"Yes")', f'=COUNTIF(\'{sh}\'!H{m["weekly_start"]}:H{m["weekly_end"]},"Yes")', f'=IFERROR(C{wsrow}/B{wsrow},0)', f'=IF(F{kpi_row}>=0.8,"Strong",IF(F{kpi_row}>=0.5,"Developing","Needs Support"))', f'=IF(OR(F{kpi_row}<0.5,E{wsrow}<0.5),"Yes — check in","No")']
        write_row(ws,n,vals,'body_center')
        apply_cell(ws.cell(n,1), STYLES['row_label'])
        apply_cell(ws.cell(n,5), STYLES['pct'])

    # status breakdown
    sb = m_header + len(intern_meta) + 3
    ws.cell(sb,1,'Overall Status Breakdown')
    style_row(ws,sb,1,3,'section')
    write_row(ws,sb+1,['Status','Count'],'table_header')
    write_row(ws,sb+2,['Completed',f'=SUM(C{start}:C{total_row-1})'],'body_left')
    write_row(ws,sb+3,['In Progress',f'=SUM(D{start}:D{total_row-1})'],'body_left')
    write_row(ws,sb+4,['Pending',f'=SUM(E{start}:E{total_row-1})'],'body_left')

    comp = sb + 6
    ws.cell(comp,1,'Completion % by Intern (data)')
    style_row(ws,comp,1,3,'section')
    write_row(ws,comp+1,['Intern','Completion %'],'table_header')
    for i,m in enumerate(intern_meta, start=comp+2):
        kpi_row = start + (i - (comp+2))
        write_row(ws,i,[m['name'],f'=F{kpi_row}'],'body_left')
        apply_cell(ws.cell(i,2), STYLES['pct'])

    weekly = comp + len(intern_meta) + 4
    ws.cell(weekly,1,'Tasks Completed per Week (per Intern)')
    style_row(ws,weekly,1,6,'section')
    write_row(ws,weekly+1,['Week']+[m['name'] for m in intern_meta],'table_header')
    max_weeks = max([m['week_count'] for m in intern_meta] + [1])
    for w in range(1, max_weeks+1):
        row = weekly + 1 + w
        vals=[f'W{w}']
        for m in intern_meta:
            vals.append(f'=COUNTIFS(\'{m["sheet"]}\'!B{m["task_start"]}:B{m["task_end"]},{w},\'{m["sheet"]}\'!E{m["task_start"]}:E{m["task_end"]},"Completed")')
        write_row(ws,row,vals,'body_center')
        apply_cell(ws.cell(row,1), STYLES['row_label'])

    # conditional formatting
    ws.conditional_formatting.add(f'G{start}:G{total_row-1}', CellIsRule(operator='equal', formula=['"On Track"'], fill=PatternFill('solid', fgColor=GREEN_FILL), font=Font(color=GREEN_TEXT)))
    ws.conditional_formatting.add(f'G{start}:G{total_row-1}', CellIsRule(operator='equal', formula=['"At Risk"'], fill=PatternFill('solid', fgColor=YELLOW_FILL), font=Font(color=YELLOW_TEXT)))
    ws.conditional_formatting.add(f'G{start}:G{total_row-1}', CellIsRule(operator='equal', formula=['"Behind"'], fill=PatternFill('solid', fgColor=RED_FILL), font=Font(color=RED_TEXT)))
    ws.conditional_formatting.add(f'F{m_header+1}:F{m_header+len(intern_meta)}', CellIsRule(operator='equal', formula=['"Strong"'], fill=PatternFill('solid', fgColor=GREEN_FILL), font=Font(color=GREEN_TEXT)))
    ws.conditional_formatting.add(f'F{m_header+1}:F{m_header+len(intern_meta)}', CellIsRule(operator='equal', formula=['"Developing"'], fill=PatternFill('solid', fgColor=YELLOW_FILL), font=Font(color=YELLOW_TEXT)))
    ws.conditional_formatting.add(f'F{m_header+1}:F{m_header+len(intern_meta)}', CellIsRule(operator='equal', formula=['"Needs Support"'], fill=PatternFill('solid', fgColor=RED_FILL), font=Font(color=RED_TEXT)))
    ws.conditional_formatting.add(f'G{m_header+1}:G{m_header+len(intern_meta)}', CellIsRule(operator='equal', formula=['"Yes — check in"'], fill=PatternFill('solid', fgColor=ORANGE_FILL), font=Font(color=YELLOW_TEXT)))

    # charts
    pie = PieChart()
    data = Reference(ws, min_col=2, min_row=sb+1, max_row=sb+4)
    labels = Reference(ws, min_col=1, min_row=sb+2, max_row=sb+4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = 'Overall Task Status'
    pie.dLbls = DataLabelList()
    pie.dLbls.showVal = True
    pie.dLbls.showCatName = True
    pie.dLbls.showPercent = True
    pie.width = 17; pie.height = 9
    ws.add_chart(pie, 'M13')

    bar = BarChart()
    data = Reference(ws, min_col=2, min_row=comp+1, max_row=comp+1+len(intern_meta))
    cats = Reference(ws, min_col=1, min_row=comp+2, max_row=comp+1+len(intern_meta))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    bar.title = 'Completion % by Intern'
    bar.dLbls = DataLabelList()
    bar.dLbls.showVal = True
    bar.dLbls.showCatName = True
    bar.dLbls.showSerName = True
    bar.y_axis.title = 'Completion %'
    bar.x_axis.title = 'Intern'
    bar.width=17; bar.height=9
    ws.add_chart(bar, 'H13')

    bar2 = BarChart()
    data = Reference(ws, min_col=2, max_col=1+len(intern_meta), min_row=weekly+1, max_row=weekly+1+max_weeks)
    cats = Reference(ws, min_col=1, min_row=weekly+2, max_row=weekly+1+max_weeks)
    bar2.add_data(data, titles_from_data=True)
    bar2.set_categories(cats)
    bar2.title = 'Tasks Completed per Week (per Intern)'
    # No data labels on weekly matrix chart because labels overlap badly when
    # many interns/weeks are present. Legend and axes are enough here.
    bar2.dLbls = None
    bar2.y_axis.title = 'Tasks Completed'
    bar2.x_axis.title = 'Week'
    bar2.width=19; bar2.height=9
    ws.add_chart(bar2, 'H28')
    for rr in range(5, weekly + max_weeks + 2):
        if rr in [5, m_header, sb+1, comp+1, weekly+1]:
            ws.row_dimensions[rr].height = 34
        elif rr in [1]:
            ws.row_dimensions[rr].height = 34
        else:
            ws.row_dimensions[rr].height = 24
    return ws
