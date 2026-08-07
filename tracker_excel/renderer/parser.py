from datetime import datetime
from dataclasses import dataclass, field
from openpyxl import load_workbook

@dataclass
class PlanSheetData:
    title: str
    subtitle: str
    headers: list
    rows: list
    sheet_name: str
    plan_type: str

@dataclass
class InternSheetData:
    name: str
    title: str
    subtitle: str
    main_headers: list
    main_row: list
    scenario_headers: list
    scenario_row: list
    task_headers: list
    tasks: list
    weekly_headers: list
    weekly_reports: list
    project_title: str
    project_headers: list
    projects: list

@dataclass
class WorkbookData:
    plans: list = field(default_factory=list)
    interns: list = field(default_factory=list)
    holidays: list = field(default_factory=list)
    versions: list = field(default_factory=list)


def row_values(ws, row, max_col):
    return [ws.cell(row, c).value for c in range(1, max_col + 1)]


def non_empty_rows(ws, start, end, max_col):
    rows = []
    for r in range(start, end + 1):
        vals = row_values(ws, r, max_col)
        if any(v is not None for v in vals):
            rows.append(vals)
    return rows


def find_row(ws, predicate):
    for r in range(1, ws.max_row + 1):
        vals = row_values(ws, r, ws.max_column)
        if predicate(vals):
            return r
    return None


def parse_plan(ws):
    title = ws['A1'].value or ws.title
    subtitle = ws['A2'].value or ''
    headers = row_values(ws, 4, ws.max_column)
    rows = non_empty_rows(ws, 5, ws.max_row, ws.max_column)
    plan_type = 'daily' if headers[:3] == ['Date', 'Day', 'Week'] else 'weekly_multitrack'
    return PlanSheetData(title, subtitle, headers, rows, ws.title, plan_type)


def parse_intern(ws):
    title = ws['A1'].value or ws.title
    subtitle = ws['A2'].value or ''
    name = ws.title
    main_header = 5
    scenario_header = 9
    task_header = 13
    weekly_header = find_row(ws, lambda vals: vals and vals[0] == 'Week #')
    project_header = find_row(ws, lambda vals: len(vals) > 1 and vals[0] == '#' and vals[1] == 'Title')
    project_section_row = None
    if project_header:
        project_section_row = project_header - 1
    task_end = (weekly_header - 3) if weekly_header else ws.max_row
    weekly_end = (project_header - 3) if project_header else ws.max_row
    return InternSheetData(
        name=name,
        title=title,
        subtitle=subtitle,
        main_headers=row_values(ws, main_header, 6),
        main_row=row_values(ws, main_header + 1, 6),
        scenario_headers=row_values(ws, scenario_header, 6),
        scenario_row=row_values(ws, scenario_header + 1, 6),
        task_headers=row_values(ws, task_header, 6),
        tasks=non_empty_rows(ws, task_header + 1, task_end, 6),
        weekly_headers=row_values(ws, weekly_header, 8) if weekly_header else [],
        weekly_reports=non_empty_rows(ws, weekly_header + 1, weekly_end, 8) if weekly_header else [],
        project_title=ws.cell(project_section_row, 1).value if project_section_row else 'SMALL PROJECTS / TASKS',
        project_headers=row_values(ws, project_header, 6) if project_header else ['#','Title','Description','Assigned Date','Due Date','Status'],
        projects=non_empty_rows(ws, project_header + 1, ws.max_row, 6) if project_header else []
    )


def parse_workbook(path):
    wb = load_workbook(path, data_only=False)
    data = WorkbookData()
    for ws in wb.worksheets:
        # Hidden history sheets round-trip by title, not by A1 content -
        # read back whatever was written on the previous render so holiday/
        # version history accumulates across renders instead of being lost
        # and re-faked from scratch each time.
        if ws.title == '_Holidays':
            data.holidays = non_empty_rows(ws, 2, ws.max_row, 5)
            continue
        if ws.title == '_Versions':
            data.versions = non_empty_rows(ws, 2, ws.max_row, 3)
            continue
        a1 = ws['A1'].value
        if ws.title == 'Dashboard' or a1 is None:
            continue
        if isinstance(a1, str) and a1.startswith('Plan '):
            data.plans.append(parse_plan(ws))
        elif isinstance(a1, str) and a1.startswith('Intern Tracker'):
            data.interns.append(parse_intern(ws))
    return data
