from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BLUE = '305496'
HEADER_BLUE = '2E75B6'
NAVY = '1F4E78'
LIGHT_GRAY = 'F2F2F2'
HOLIDAY_YELLOW = 'FFE699'
DARK_GRAY = '404040'
WHITE = 'FFFFFF'
BLACK = '000000'
GREEN_FILL = 'C6EFCE'
GREEN_TEXT = '006100'
RED_FILL = 'FFC7CE'
RED_TEXT = '9C0006'
YELLOW_FILL = 'FFEB9C'
YELLOW_TEXT = '9C6500'
ORANGE_FILL = 'FCE4D6'

thin = Side(style='thin', color='808080')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def font(size=10, bold=False, color=BLACK, name='Arial'):
    return Font(name=name, size=size, bold=bold, color=color)


def fill(color):
    return PatternFill('solid', fgColor=color)


def align(horizontal='center', vertical='center', wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


def apply_cell(cell, style):
    if 'font' in style: cell.font = style['font']
    if 'fill' in style: cell.fill = style['fill']
    if 'border' in style: cell.border = style['border']
    if 'alignment' in style: cell.alignment = style['alignment']
    if 'num_format' in style: cell.number_format = style['num_format']


def apply_range(ws, cell_range, style):
    for row in ws[cell_range]:
        for cell in row:
            apply_cell(cell, style)


STYLES = {
    'dashboard_title': {'font': font(18, True, WHITE), 'fill': fill(BLUE), 'alignment': align('center', 'center', False)},
    'intern_title': {'font': font(16, True, WHITE), 'fill': fill(BLUE), 'alignment': align('center', 'center', False)},
    'plan_title': {'font': font(16, True, NAVY), 'alignment': Alignment(horizontal='left')},
    'subtitle': {'font': font(10, False, DARK_GRAY), 'alignment': Alignment(horizontal='left')},
    'subtitle_center': {'font': font(10, False, DARK_GRAY), 'alignment': Alignment(horizontal='center')},
    'section': {'font': font(11, True, WHITE), 'fill': fill(BLUE), 'border': BORDER, 'alignment': align('center')},
    'table_header': {'font': font(11, True, WHITE), 'fill': fill(HEADER_BLUE), 'border': BORDER, 'alignment': align('center')},
    'plan_header': {'font': font(11, True, WHITE), 'fill': fill(NAVY), 'border': BORDER, 'alignment': align('center')},
    'body_center': {'font': font(10), 'border': BORDER, 'alignment': align('center')},
    'body_left': {'font': font(10), 'border': BORDER, 'alignment': align('left')},
    'body_center_gray': {'font': font(10), 'fill': fill(LIGHT_GRAY), 'border': BORDER, 'alignment': align('center')},
    'body_left_gray': {'font': font(10), 'fill': fill(LIGHT_GRAY), 'border': BORDER, 'alignment': align('left')},
    'row_label': {'font': font(10, True), 'border': BORDER, 'alignment': align('left')},
    'pct': {'font': font(10), 'border': BORDER, 'alignment': align('center'), 'num_format': '0.0%'},
    'date': {'font': font(10), 'border': BORDER, 'alignment': align('center'), 'num_format': 'ddd, dd-mmm-yyyy'},
    'short_date': {'font': font(10), 'border': BORDER, 'alignment': align('center'), 'num_format': 'dd-mmm-yyyy'},
    'holiday_date': {'font': font(10), 'fill': fill(HOLIDAY_YELLOW), 'border': BORDER, 'alignment': align('center'), 'num_format': 'ddd, dd-mmm-yyyy'},
    'holiday_center': {'font': font(10), 'fill': fill(HOLIDAY_YELLOW), 'border': BORDER, 'alignment': align('center')},
    'holiday_left': {'font': font(10), 'fill': fill(HOLIDAY_YELLOW), 'border': BORDER, 'alignment': align('left')},
}
