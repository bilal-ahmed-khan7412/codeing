from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font
from .styles import STYLES, apply_cell, apply_range, GREEN_FILL, GREEN_TEXT, RED_FILL, RED_TEXT, YELLOW_FILL, YELLOW_TEXT
from .utils import default_visible_page, style_row, write_row, set_widths, is_date, max_numeric_week, set_dynamic_row_height


def render_intern(wb, intern):
    ws = wb.create_sheet(intern.name[:31])
    default_visible_page(ws)
    set_widths(ws, {'A':34,'B':48,'C':38,'D':82,'E':27,'F':76,'G':16,'H':22})
    width_map = {1:34,2:48,3:38,4:82,5:27,6:76,7:16,8:22}
    section_end_col = 8
    # row 1/2 title strips, no merged cells
    row = 1
    ws.cell(row,1,intern.title)
    style_row(ws,row,1,section_end_col,'intern_title')
    ws.row_dimensions[row].height = 24
    row += 1
    ws.cell(row,1,intern.subtitle)
    style_row(ws,row,1,section_end_col,'subtitle')
    row += 2

    # main project
    ws.cell(row,1,'MAIN PROJECT')
    style_row(ws,row,1,6,'section')
    row += 1
    write_row(ws,row,intern.main_headers,'table_header')
    row += 1
    write_row(ws,row,intern.main_row,'body_left')
    for c in [4,5]:
        if is_date(ws.cell(row,c).value):
            apply_cell(ws.cell(row,c), STYLES['short_date'])
    apply_cell(ws.cell(row,6), STYLES['body_center'])
    set_dynamic_row_height(ws, row, intern.main_row, width_map, min_height=90, max_height=190)
    row += 2

    # scenario
    ws.cell(row,1,'REAL-WORLD SCENARIO')
    style_row(ws,row,1,6,'section')
    row += 1
    write_row(ws,row,intern.scenario_headers,'table_header')
    row += 1
    write_row(ws,row,intern.scenario_row,'body_left')
    for c in [4,5,6]:
        apply_cell(ws.cell(row,c), STYLES['body_center'])
    if is_date(ws.cell(row,5).value):
        apply_cell(ws.cell(row,5), STYLES['short_date'])
    set_dynamic_row_height(ws, row, intern.scenario_row, width_map, min_height=90, max_height=190)
    row += 2

    # daily tasks
    ws.cell(row,1,'DAILY TASKS')
    style_row(ws,row,1,6,'section')
    daily_section_row = row
    row += 1
    write_row(ws,row,intern.task_headers,'table_header')
    task_header_row = row
    row += 1
    task_start = row
    for i, vals in enumerate(intern.tasks):
        is_holiday = any(isinstance(x,str) and ('HOLIDAY' in x.upper() or 'Holiday' == str(x)) for x in vals)
        style_center = 'holiday_center' if is_holiday else ('body_center_gray' if i % 2 else 'body_center')
        style_left = 'holiday_left' if is_holiday else ('body_left_gray' if i % 2 else 'body_left')
        for c, val in enumerate(vals, 1):
            ws.cell(row,c,val)
            if c == 1 and is_date(val):
                apply_cell(ws.cell(row,c), STYLES['holiday_date'] if is_holiday else STYLES['date'])
            elif c in [3,4,6]:
                apply_cell(ws.cell(row,c), STYLES[style_left])
            else:
                apply_cell(ws.cell(row,c), STYLES[style_center])
        row += 1
        set_dynamic_row_height(ws, row-1, vals, width_map, min_height=38, max_height=140)
    task_end = row - 1
    row += 1

    # weekly updates
    ws.cell(row,1,'WEEKLY UPDATES')
    style_row(ws,row,1,8,'section')
    weekly_section_row = row
    row += 1
    weekly_header_row = row
    write_row(ws,row,intern.weekly_headers,'table_header')
    row += 1
    weekly_start = row
    week_count = max([len(intern.weekly_reports), max_numeric_week(intern.tasks), 1])
    for idx in range(week_count):
        source = intern.weekly_reports[idx] if idx < len(intern.weekly_reports) else [idx+1,'','','','','','No','No']
        values = list(source) + [''] * (8 - len(source))
        values[0] = idx + 1 if not values[0] else values[0]
        # E is calculated completed task count
        values[4] = f'=COUNTIFS(B{task_start}:B{task_end},{idx+1},E{task_start}:E{task_end},"Completed")'
        write_row(ws,row,values,'body_left')
        for c in [1,5,7,8]:
            apply_cell(ws.cell(row,c), STYLES['body_center'])
        row += 1
        set_dynamic_row_height(ws, row-1, values, width_map, min_height=40, max_height=160)
    weekly_end = row - 1
    row += 1

    # small projects
    project_section_row = row
    ws.cell(row,1,intern.project_title or 'SMALL PROJECTS / TASKS')
    style_row(ws,row,1,6,'section')
    row += 1
    write_row(ws,row,intern.project_headers,'table_header')
    row += 1
    project_start = row
    for vals in intern.projects:
        write_row(ws,row,vals,'body_left')
        for c in [1,4,5,6]:
            apply_cell(ws.cell(row,c), STYLES['body_center'])
        for c in [4,5]:
            if is_date(ws.cell(row,c).value):
                apply_cell(ws.cell(row,c), STYLES['short_date'])
        row += 1
        set_dynamic_row_height(ws, row-1, vals, width_map, min_height=42, max_height=150)
    project_end = row - 1 if intern.projects else project_start - 1

    # section and header heights
    for rr in [1,2,daily_section_row,task_header_row,weekly_section_row,weekly_header_row,project_section_row,project_section_row+1]:
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height or 0, 26)
    for rr in [4,5,8,9,12,13]:
        ws.row_dimensions[rr].height = max(ws.row_dimensions[rr].height or 0, 28)

    # validations
    dv_status = DataValidation(type='list', formula1='"Pending,In Progress,Completed"', allow_blank=True)
    dv_yes = DataValidation(type='list', formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_status)
    ws.add_data_validation(dv_yes)
    if task_end >= task_start:
        dv_status.add(f'E{task_start}:E{task_end}')
        # conditional status formatting
        ws.conditional_formatting.add(f'E{task_start}:E{task_end}', CellIsRule(operator='equal', formula=['"Completed"'], fill=PatternFill('solid', fgColor=GREEN_FILL), font=Font(color=GREEN_TEXT)))
        ws.conditional_formatting.add(f'E{task_start}:E{task_end}', CellIsRule(operator='equal', formula=['"Pending"'], fill=PatternFill('solid', fgColor=RED_FILL), font=Font(color=RED_TEXT)))
        ws.conditional_formatting.add(f'E{task_start}:E{task_end}', CellIsRule(operator='equal', formula=['"In Progress"'], fill=PatternFill('solid', fgColor=YELLOW_FILL), font=Font(color=YELLOW_TEXT)))
    if weekly_end >= weekly_start:
        dv_yes.add(f'G{weekly_start}:H{weekly_end}')
        ws.conditional_formatting.add(f'G{weekly_start}:H{weekly_end}', CellIsRule(operator='equal', formula=['"Yes"'], fill=PatternFill('solid', fgColor=GREEN_FILL), font=Font(color=GREEN_TEXT)))
    if project_end >= project_start:
        dv_status.add(f'F{project_start}:F{project_end}')
        ws.conditional_formatting.add(f'F{project_start}:F{project_end}', CellIsRule(operator='equal', formula=['"Completed"'], fill=PatternFill('solid', fgColor=GREEN_FILL), font=Font(color=GREEN_TEXT)))
        ws.conditional_formatting.add(f'F{project_start}:F{project_end}', CellIsRule(operator='equal', formula=['"Pending"'], fill=PatternFill('solid', fgColor=RED_FILL), font=Font(color=RED_TEXT)))
        ws.conditional_formatting.add(f'F{project_start}:F{project_end}', CellIsRule(operator='equal', formula=['"In Progress"'], fill=PatternFill('solid', fgColor=YELLOW_FILL), font=Font(color=YELLOW_TEXT)))

    return {
        'name': intern.name,
        'sheet': ws.title,
        'task_start': task_start,
        'task_end': task_end,
        'weekly_start': weekly_start,
        'weekly_end': weekly_end,
        'project_start': project_start,
        'project_end': project_end,
        'week_count': week_count,
        'target_end': intern.main_row[4] if len(intern.main_row) >= 5 else None,
    }
