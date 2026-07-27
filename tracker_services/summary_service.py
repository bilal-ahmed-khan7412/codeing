from openpyxl import load_workbook
from tracker_core.models import CommandResult

class SummaryService:
    def generate_progress_summary(self, workbook_path: str, intern_name: str | None = None) -> CommandResult:
        wb = load_workbook(workbook_path, data_only=False)
        lines = []
        target_sheets = []
        for ws in wb.worksheets:
            a1 = ws['A1'].value
            if isinstance(a1, str) and a1.startswith('Intern Tracker'):
                if intern_name is None or ws.title.lower() == intern_name.lower():
                    target_sheets.append(ws)
        for ws in target_sheets:
            # find daily task and weekly sections
            daily = weekly = small = None
            for r in range(1, ws.max_row + 1):
                v = ws.cell(r,1).value
                if v == 'DAILY TASKS': daily = r
                elif v == 'WEEKLY UPDATES': weekly = r
                elif isinstance(v,str) and v.startswith('SMALL PROJECTS'): small = r
            if not daily or not weekly:
                continue
            task_start = daily + 2
            task_end = weekly - 2
            statuses = [ws.cell(r,5).value for r in range(task_start, task_end+1) if ws.cell(r,5).value]
            total = len(statuses)
            completed = statuses.count('Completed')
            in_progress = statuses.count('In Progress')
            pending = statuses.count('Pending')
            pct = completed / total if total else 0
            lines.append(f"Intern: {ws.title}")
            lines.append(f"Total Tasks: {total}")
            lines.append(f"Completed: {completed}")
            lines.append(f"In Progress: {in_progress}")
            lines.append(f"Pending: {pending}")
            lines.append(f"Completion: {pct:.1%}")
            lines.append("")
        return CommandResult(True, "Generated progress summary", data={"summary": "\n".join(lines)})
