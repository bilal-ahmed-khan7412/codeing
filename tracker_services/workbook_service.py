from tracker_core.models import CommandResult
from tracker_services.render_service import RenderService
from tracker_services.version_service import VersionService

class WorkbookService:

    def create_fresh_workbook(self, output_path: str) -> CommandResult:
        """Create a fresh blank workbook with dashboard shell and hidden system sheets.

        This does not require a source workbook. It creates an empty automation-ready
        tracker that can later receive interns/plans using commands.
        """
        from openpyxl import Workbook
        from tracker_excel.renderer.hidden_renderer import render_hidden_sheets
        from tracker_excel.renderer.parser import WorkbookData
        from tracker_excel.renderer.styles import STYLES, apply_cell
        from tracker_excel.renderer.utils import default_visible_page, set_widths, style_row, write_row

        data = WorkbookData()
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet('Dashboard', 0)
        default_visible_page(ws)
        set_widths(ws, {'A':30,'B':16,'C':16,'D':16,'E':16,'F':18,'G':18,'H':18,'I':22,'J':48})
        ws.cell(1,1,'Internship Learning — Dashboard - Systems Limited')
        style_row(ws,1,1,10,'dashboard_title')
        ws.row_dimensions[1].height = 32
        ws.cell(2,1,'No interns added yet. Use add-intern to start tracking.')
        style_row(ws,2,1,10,'subtitle_center')

        ws.cell(4,1,'KPIs — Per Intern')
        style_row(ws,4,1,10,'section')
        write_row(ws,5,['Intern','Total Tasks','Completed','In Progress','Pending','Completion %','On Track?','Current Week','Weekly Projects Done','Days Left'],'table_header')
        write_row(ws,6,['No interns yet','','','','','','','','',''],'body_center')

        ws.cell(9,1,'Manager View — Reporting & Engagement')
        style_row(ws,9,1,7,'section')
        write_row(ws,10,['Intern','Weeks in Plan','Weekly Emails Sent','LM Acknowledged','Reporting %','Learning Health','Attention Needed?'],'table_header')
        write_row(ws,11,['No interns yet','','','','','',''],'body_center')

        ws.cell(14,1,'Overall Status Breakdown')
        style_row(ws,14,1,3,'section')
        write_row(ws,15,['Status','Count'],'table_header')
        write_row(ws,16,['Completed',0],'body_left')
        write_row(ws,17,['In Progress',0],'body_left')
        write_row(ws,18,['Pending',0],'body_left')

        ws.cell(20,1,'Completion % by Intern (data)')
        style_row(ws,20,1,3,'section')
        write_row(ws,21,['Intern','Completion %'],'table_header')
        write_row(ws,22,['No interns yet',0],'body_left')
        apply_cell(ws.cell(22,2), STYLES['pct'])

        ws.cell(25,1,'Tasks Completed per Week (per Intern)')
        style_row(ws,25,1,6,'section')
        write_row(ws,26,['Week'],'table_header')
        for r in range(27,35):
            write_row(ws,r,[f'W{r-26}'],'body_center')

        render_hidden_sheets(wb, data, [])
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.save(output_path)
        return CommandResult(True, f"Created fresh blank workbook: {output_path}", output_path)

    def create_clean_version(self, source_path: str, output_path: str | None = None) -> CommandResult:
        out = output_path or VersionService.next_version_path(source_path)
        RenderService.render_from_workbook(source_path, out)
        return CommandResult(True, f"Created clean rendered workbook: {out}", out)
