
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any
import json
import uuid

from openpyxl import load_workbook

try:
    from tracker_excel.renderer.parser import parse_workbook
except Exception:
    parse_workbook = None

try:
    from tracker_config.settings import load_settings
    from tracker_llm.providers import build_provider
except Exception:
    load_settings = None
    build_provider = None

BASE_DIR = Path(__file__).resolve().parents[1]
UPLOAD_DIR = BASE_DIR / 'uploads'
OUTPUT_DIR = BASE_DIR / 'outputs'
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)

SUBJECTIVE_CRITERIA = [
    'Skills Acquired (plan themes)',
    'Final / Main Project',
    'Real-World Scenario Support',
    'Meets Deadlines',
    'Problem Solving',
    'Creative Thinking',
    'Out-of-the-Box Thinking',
    'Thinking Structure (reasoning)',
    'Decision Making',
    'Communication',
    'Group Participation / Teamwork',
    'Compliance (security/process/reporting)',
    'Ownership & Initiative',
    'Learning Agility / Self-Learning',
    'Documentation Quality',
    'Reliability & Consistency',
]

RUBRICS = {
    'Skills Acquired (plan themes)': '0=not started, 3=works with help, 5=independent/expert across plan skill areas.',
    'Final / Main Project': '0=not attempted, 2=partial/heavy help, 3=working/meets scope, 4=solid/minor help, 5=excellent demo-ready.',
    'Real-World Scenario Support': '0=no contribution, 2=observed only, 3=handled with guidance, 4=mostly independent, 5=led/resolved with strong RCA.',
    'Meets Deadlines': '0=chronically late, 2=often slips, 3=mostly on time, 4=reliably on time, 5=consistently early/on time.',
    'Problem Solving': '0=cannot progress alone, 2=needs constant help, 3=solves routine issues, 4=solves most alone, 5=diagnoses complex problems independently.',
    'Creative Thinking': '0=none, 2=rarely, 3=some original ideas, 4=frequently, 5=consistently inventive solutions.',
    'Out-of-the-Box Thinking': '0=strictly literal, 2=seldom, 3=occasional novel angle, 4=often reframes, 5=non-obvious approaches.',
    'Thinking Structure (reasoning)': '0=disorganised, 2=scattered, 3=logical with prompting, 4=clear steps, 5=structured first-principles reasoning.',
    'Decision Making': '0=avoids/guesses, 2=hesitant, 3=reasonable with data, 4=weighs options, 5=sound judgment on trade-offs.',
    'Communication': '0=unclear/absent, 2=minimal, 3=clear routine updates, 4=proactive, 5=articulate and audience-aware.',
    'Group Participation / Teamwork': '0=disengaged, 2=passive, 3=cooperative, 4=active contributor, 5=elevates team/helps peers.',
    'Compliance (security/process/reporting)': '0=ignores policy, 2=frequent lapses, 3=follows with reminders, 4=disciplined, 5=audit-ready/reports weekly.',
    'Ownership & Initiative': '0=waits to be told, 2=minimal, 3=owns assigned work, 4=anticipates, 5=proactively takes on more.',
    'Learning Agility / Self-Learning': '0=struggles with new topics, 2=slow, 3=learns with support, 4=self-directed, 5=rapid independent learner.',
    'Documentation Quality': '0=none, 2=sparse, 3=usable notes, 4=clear, 5=reusable docs/runbooks.',
    'Reliability & Consistency': '0=unreliable, 2=uneven, 3=steady, 4=dependable, 5=consistent and dependable throughout.',
}

QUESTIONS = {
    'Skills Acquired (plan themes)': 'Across the plan skill areas, what proficiency did the intern demonstrate? Mention independent areas and areas needing help.',
    'Final / Main Project': 'How complete and demo-ready was the final/main project, and how much guidance was required?',
    'Real-World Scenario Support': 'How did the intern contribute to the real-world scenario? Did the intern observe, assist, handle with guidance, or lead?',
    'Meets Deadlines': 'Did the intern complete assigned daily work and weekly projects on time? Mention delays or consistency.',
    'Problem Solving': 'When blocked, how independently did the intern diagnose and solve issues?',
    'Creative Thinking': 'Did the intern suggest original ideas, improvements, or alternative implementations?',
    'Out-of-the-Box Thinking': 'Did the intern reframe problems or suggest non-obvious approaches?',
    'Thinking Structure (reasoning)': 'How structured was the intern’s reasoning and explanation of steps?',
    'Decision Making': 'How well did the intern make trade-off decisions using available data?',
    'Communication': 'How clear, timely, and proactive were the intern’s updates?',
    'Group Participation / Teamwork': 'How well did the intern participate with the team and help peers?',
    'Compliance (security/process/reporting)': 'How consistently did the intern follow process, security rules, reporting, and documentation expectations?',
    'Ownership & Initiative': 'Did the intern own assigned work and proactively take next steps?',
    'Learning Agility / Self-Learning': 'How quickly did the intern learn new topics and self-study?',
    'Documentation Quality': 'How useful and reusable were the intern’s notes, reports, and documentation?',
    'Reliability & Consistency': 'How dependable and consistent was the intern throughout the evaluation period?',
}


def save_upload(upload_file, prefix: str) -> Path:
    suffix = Path(upload_file.filename or 'upload.xlsx').suffix or '.xlsx'
    safe_name = f"{prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}{suffix}"
    path = UPLOAD_DIR / safe_name
    with path.open('wb') as f:
        f.write(upload_file.file.read())
    return path


def normalize_name(name: str) -> str:
    return ''.join(ch.lower() for ch in (name or '') if ch.isalnum() or ch.isspace()).strip()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize_name(a), normalize_name(b)).ratio()


def get_tracker_interns(tracker_path: str) -> list[dict[str, Any]]:
    if parse_workbook is None:
        return []
    data = parse_workbook(tracker_path)
    interns = []
    for item in getattr(data, 'interns', []) or []:
        start = item.main_row[3] if len(item.main_row) > 3 else ''
        end = item.main_row[4] if len(item.main_row) > 4 else ''
        interns.append({
            'name': item.name,
            'sheet': getattr(item, 'sheet_name', item.name),
            'plan': getattr(item, 'plan_name', '') or '',
            'start': start.strftime('%Y-%m-%d') if hasattr(start, 'strftime') else str(start or ''),
            'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else str(end or ''),
        })
    return interns


def get_eval_scorecards(eval_path: str) -> list[dict[str, Any]]:
    wb = load_workbook(eval_path, data_only=False)
    cards = []
    for ws in wb.worksheets:
        if ws.title.startswith('SC - '):
            cards.append({'name': ws.title.replace('SC - ', '').strip(), 'sheet': ws.title})
    return cards


def match_candidates(tracker_intern: str, eval_cards: list[dict[str, Any]], limit: int = 5) -> list[dict[str, Any]]:
    scored = []
    for card in eval_cards:
        score = similarity(tracker_intern, card['name'])
        scored.append({**card, 'match': round(score * 100, 1)})
    return sorted(scored, key=lambda x: x['match'], reverse=True)[:limit]


def status_done(value: Any) -> bool:
    return str(value or '').strip().lower() in {'completed', 'complete', 'done', 'closed'}


def band_score(completed: int, planned: int) -> int:
    if planned <= 0 or completed <= 0:
        return 0
    pct = completed / planned
    if pct < 0.20:
        return 1
    if pct < 0.50:
        return 2
    if pct < 0.70:
        return 3
    if pct < 0.90:
        return 4
    return 5


def get_tracker_metrics(tracker_path: str, intern_name: str) -> dict[str, Any]:
    if parse_workbook is None:
        return {}
    data = parse_workbook(tracker_path)
    intern = None
    for item in getattr(data, 'interns', []) or []:
        if normalize_name(item.name) == normalize_name(intern_name):
            intern = item
            break
    if not intern:
        # closest tracker intern fallback
        best = None
        best_score = 0
        for item in getattr(data, 'interns', []) or []:
            sc = similarity(item.name, intern_name)
            if sc > best_score:
                best = item; best_score = sc
        intern = best
    if not intern:
        return {}

    tasks = getattr(intern, 'tasks', []) or []
    projects = getattr(intern, 'projects', []) or []
    daily_planned = len(tasks)
    daily_done = sum(1 for row in tasks if len(row) > 4 and status_done(row[4]))
    weekly_planned = len(projects)
    weekly_done = sum(1 for row in projects if len(row) > 5 and status_done(row[5]))

    start = intern.main_row[3] if len(intern.main_row) > 3 else ''
    end = intern.main_row[4] if len(intern.main_row) > 4 else ''
    plan = getattr(intern, 'plan_name', '') or ''
    main_project = intern.main_row[0] if len(intern.main_row) > 0 else ''
    scenario = intern.scenario_row[0] if len(intern.scenario_row) > 0 else ''

    return {
        'matched_tracker_name': intern.name,
        'plan': plan,
        'start': start.strftime('%Y-%m-%d') if hasattr(start, 'strftime') else str(start or ''),
        'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else str(end or ''),
        'daily_done': daily_done,
        'daily_planned': daily_planned,
        'daily_pct': daily_done / daily_planned if daily_planned else 0,
        'weekly_done': weekly_done,
        'weekly_planned': weekly_planned,
        'weekly_pct': weekly_done / weekly_planned if weekly_planned else 0,
        'daily_score': band_score(daily_done, daily_planned),
        'weekly_score': band_score(weekly_done, weekly_planned),
        'main_project': str(main_project or ''),
        'scenario': str(scenario or ''),
    }


def build_questions() -> list[dict[str, Any]]:
    return [
        {
            'criterion': c,
            'question': QUESTIONS[c],
            'rubric': RUBRICS[c],
            'weight_note': 'Score 0–5. Admin can edit final score before writing.'
        }
        for c in SUBJECTIVE_CRITERIA
    ]


def _llm_provider():
    if not load_settings or not build_provider:
        return None
    try:
        settings = load_settings('.env')
        if settings.ai_provider.lower() == 'mock':
            return None
        return build_provider(settings)
    except Exception:
        return None


def heuristic_score(answer: str) -> int:
    a = (answer or '').lower()
    if any(w in a for w in ['outstanding', 'excellent', 'independent', 'led', 'demo-ready', 'demo ready', 'expert']):
        return 5
    if any(w in a for w in ['exceeds', 'mostly independent', 'proactive', 'solid', 'minor help', 'reliable']):
        return 4
    if any(w in a for w in ['meets', 'with guidance', 'routine', 'average', 'mostly on time', 'working']):
        return 3
    if any(w in a for w in ['partial', 'needs help', 'needs constant', 'often', 'below', 'observed']):
        return 2
    if any(w in a for w in ['poor', 'rarely', 'minimal', 'late', 'struggled']):
        return 1
    return 3 if answer.strip() else 0


def suggest_score(criterion: str, answer: str, tracker_context: dict[str, Any]) -> dict[str, Any]:
    provider = _llm_provider()
    rubric = RUBRICS.get(criterion, '')
    if provider:
        prompt = f'''
You are assisting an admin with an intern evaluation. Return ONLY JSON.
Criterion: {criterion}
Rubric: {rubric}
Tracker context: {json.dumps(tracker_context, default=str)}
Evaluator answer: {answer}

Return shape:
{{"score": 0-5, "rationale": "brief specific rationale based only on the answer and tracker context"}}
'''
        try:
            data = provider.complete_json(prompt)
            score = int(data.get('score', heuristic_score(answer)))
            score = max(0, min(5, score))
            return {'score': score, 'rationale': str(data.get('rationale', '')).strip() or 'Suggested from evaluator answer.'}
        except Exception:
            pass
    score = heuristic_score(answer)
    return {'score': score, 'rationale': 'Suggested from evaluator answer using rubric keywords. Admin should review/edit.'}


def _find_cell(ws, text: str):
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value or '').strip() == text:
                return cell
    return None


def _write_below_label(ws, label: str, value: Any):
    cell = _find_cell(ws, label)
    if cell:
        ws.cell(row=cell.row + 1, column=cell.column).value = value


def _write_after_label(ws, label: str, value: Any, offset: int = 1):
    cell = _find_cell(ws, label)
    if cell:
        ws.cell(row=cell.row, column=cell.column + offset).value = value


def _write_criterion(ws, criterion: str, score: Any, comment: str = ''):
    cell = _find_cell(ws, criterion)
    if not cell:
        return
    # Expected layout: Criterion in column B, Score in D, Evidence/comments in G.
    ws.cell(row=cell.row, column=cell.column + 2).value = int(score)
    if comment:
        ws.cell(row=cell.row, column=cell.column + 5).value = comment


def finalize_evaluation(eval_path: str, eval_sheet: str, tracker_metrics: dict[str, Any], scores: dict[str, Any], comments: dict[str, str], strengths: str, development: str, remark: str) -> Path:
    wb = load_workbook(eval_path)
    if eval_sheet not in wb.sheetnames:
        raise ValueError(f'Evaluation sheet not found: {eval_sheet}')
    ws = wb[eval_sheet]

    # Delivery snapshot fields already exist in workbook. Do not add fields.
    _write_below_label(ws, 'Daily tasks completed', tracker_metrics.get('daily_done', 0))
    _write_below_label(ws, 'Daily tasks planned', tracker_metrics.get('daily_planned', 0))
    _write_below_label(ws, 'Daily completion %', tracker_metrics.get('daily_pct', 0))
    _write_below_label(ws, 'Weekly projects done', tracker_metrics.get('weekly_done', 0))
    _write_below_label(ws, 'Weekly projects planned', tracker_metrics.get('weekly_planned', 0))

    # Auto scores for delivery completion criteria.
    _write_criterion(ws, 'Daily Tasks Completion', tracker_metrics.get('daily_score', 0), 'Auto-filled from tracker daily task completion.')
    _write_criterion(ws, 'Weekly Tasks / Projects', tracker_metrics.get('weekly_score', 0), 'Auto-filled from tracker weekly project completion.')

    # Admin-approved scores from interview questions.
    for c, score in scores.items():
        _write_criterion(ws, c, score, comments.get(c, ''))

    # Existing text areas.
    _write_below_label(ws, 'Key strengths', strengths)
    _write_below_label(ws, 'Development areas', development)
    _write_below_label(ws, 'Final manager remark', remark)

    out = OUTPUT_DIR / f"Evaluated_{Path(eval_path).stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(out)
    return out


# v0.72 evaluation rationale override
# Each answer should produce a score plus a clear reason for that score.
# If the admin edits the score, the UI marks it as "Set by user". Rationale is
# page-only and is not written to the evaluation workbook.
def _v72_compact_answer(answer: str, limit: int = 160) -> str:
    answer = ' '.join(str(answer or '').split())
    if len(answer) > limit:
        return answer[:limit].rstrip() + '...'
    return answer


def _v72_reason_for_score(criterion: str, answer: str, score: int, tracker_context: dict[str, Any]) -> str:
    ans = _v72_compact_answer(answer)
    if not ans:
        return f'No evaluator evidence was provided for {criterion}; score {score}/5 should be reviewed by the admin.'
    if score >= 5:
        return f'Score 5/5 because the evaluator answer indicates outstanding or independent performance for {criterion}: "{ans}".'
    if score == 4:
        return f'Score 4/5 because the evaluator answer indicates strong performance that exceeds expectations for {criterion}, but admin should confirm evidence before finalizing: "{ans}".'
    if score == 3:
        return f'Score 3/5 because the evaluator answer suggests the intern meets expectations for {criterion}, with no clear evidence for a higher score: "{ans}".'
    if score == 2:
        return f'Score 2/5 because the evaluator answer suggests below-expectation performance or notable support required for {criterion}: "{ans}".'
    if score == 1:
        return f'Score 1/5 because the evaluator answer suggests poor or minimal demonstration for {criterion}: "{ans}".'
    return f'Score 0/5 because there is no demonstrated evidence for {criterion} in the evaluator answer.'


def suggest_score(criterion: str, answer: str, tracker_context: dict[str, Any]) -> dict[str, Any]:
    provider = _llm_provider()
    rubric = RUBRICS.get(criterion, '')
    if provider:
        prompt = f'''
You are assisting an admin with an intern evaluation. Return ONLY JSON.
For this ONE criterion, suggest a score and explain why.
Criterion: {criterion}
Rubric: {rubric}
Tracker context: {json.dumps(tracker_context, default=str)}
Evaluator answer: {answer}

Rules:
- Score must be an integer from 0 to 5.
- Rationale must explain why this score was chosen for this criterion.
- Do not say only "Suggested from evaluator answer".
- If answer is vague, mention that evidence is limited and admin should confirm/edit.

Return shape:
{{"score": 0-5, "rationale": "specific reason for this score"}}
'''
        try:
            data = provider.complete_json(prompt)
            score = int(data.get('score', heuristic_score(answer)))
            score = max(0, min(5, score))
            rationale = str(data.get('rationale', '')).strip()
            if not rationale or rationale.lower() in {'suggested from evaluator answer.', 'suggested from evaluator answer'}:
                rationale = _v72_reason_for_score(criterion, answer, score, tracker_context)
            return {'score': score, 'rationale': rationale, 'source': 'AI suggested'}
        except Exception:
            pass
    score = heuristic_score(answer)
    rationale = _v72_reason_for_score(criterion, answer, score, tracker_context)
    return {'score': score, 'rationale': rationale, 'source': 'AI suggested'}


# v0.73 evaluation basis override
# Supports two fair scoring bases:
# - as_of: count only tasks/projects due up to evaluation_date
# - overall: count full internship tasks/projects

def _v73_parse_date(value):
    if value is None or value == '':
        return None
    if hasattr(value, 'date'):
        return value
    try:
        return datetime.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _v73_row_date(row, idx):
    try:
        return _v73_parse_date(row[idx])
    except Exception:
        return None


def _v73_due_filter(date_value, evaluation_dt, basis: str) -> bool:
    if basis == 'overall' or evaluation_dt is None:
        return True
    if date_value is None:
        return False
    return date_value.date() <= evaluation_dt.date()


def _v73_pct(done: int, planned: int) -> float:
    return done / planned if planned else 0


def get_tracker_metrics(tracker_path: str, intern_name: str, evaluation_date: str | None = None, basis: str = 'as_of') -> dict[str, Any]:
    if parse_workbook is None:
        return {}
    basis = (basis or 'as_of').lower()
    if basis not in {'as_of', 'overall'}:
        basis = 'as_of'
    evaluation_dt = _v73_parse_date(evaluation_date) or datetime.now()

    data = parse_workbook(tracker_path)
    intern = None
    for item in getattr(data, 'interns', []) or []:
        if normalize_name(item.name) == normalize_name(intern_name):
            intern = item
            break
    if not intern:
        best = None
        best_score = 0
        for item in getattr(data, 'interns', []) or []:
            sc = similarity(item.name, intern_name)
            if sc > best_score:
                best = item
                best_score = sc
        intern = best
    if not intern:
        return {}

    tasks = getattr(intern, 'tasks', []) or []
    projects = getattr(intern, 'projects', []) or []

    overall_daily_planned = len(tasks)
    overall_daily_done = sum(1 for row in tasks if len(row) > 4 and status_done(row[4]))
    overall_weekly_planned = len(projects)
    overall_weekly_done = sum(1 for row in projects if len(row) > 5 and status_done(row[5]))

    asof_tasks = [row for row in tasks if _v73_due_filter(_v73_row_date(row, 0), evaluation_dt, 'as_of')]
    asof_projects = []
    for row in projects:
        # Project layout is usually [#, title, description, assigned_date, due_date, status].
        due = _v73_row_date(row, 4)
        if due is None:
            due = _v73_row_date(row, 3)
        if _v73_due_filter(due, evaluation_dt, 'as_of'):
            asof_projects.append(row)

    asof_daily_planned = len(asof_tasks)
    asof_daily_done = sum(1 for row in asof_tasks if len(row) > 4 and status_done(row[4]))
    asof_weekly_planned = len(asof_projects)
    asof_weekly_done = sum(1 for row in asof_projects if len(row) > 5 and status_done(row[5]))

    if basis == 'overall':
        daily_planned = overall_daily_planned
        daily_done = overall_daily_done
        weekly_planned = overall_weekly_planned
        weekly_done = overall_weekly_done
    else:
        daily_planned = asof_daily_planned
        daily_done = asof_daily_done
        weekly_planned = asof_weekly_planned
        weekly_done = asof_weekly_done

    start = intern.main_row[3] if len(intern.main_row) > 3 else ''
    end = intern.main_row[4] if len(intern.main_row) > 4 else ''
    plan = getattr(intern, 'plan_name', '') or ''
    main_project = intern.main_row[0] if len(intern.main_row) > 0 else ''
    scenario = intern.scenario_row[0] if len(intern.scenario_row) > 0 else ''

    return {
        'matched_tracker_name': intern.name,
        'plan': plan,
        'start': start.strftime('%Y-%m-%d') if hasattr(start, 'strftime') else str(start or ''),
        'end': end.strftime('%Y-%m-%d') if hasattr(end, 'strftime') else str(end or ''),
        'evaluation_date': evaluation_dt.strftime('%Y-%m-%d'),
        'basis': basis,

        # Selected basis values used for scoring/writing.
        'daily_done': daily_done,
        'daily_planned': daily_planned,
        'daily_pct': _v73_pct(daily_done, daily_planned),
        'weekly_done': weekly_done,
        'weekly_planned': weekly_planned,
        'weekly_pct': _v73_pct(weekly_done, weekly_planned),
        'daily_score': band_score(daily_done, daily_planned),
        'weekly_score': band_score(weekly_done, weekly_planned),

        # Diagnostic comparison values displayed only on page.
        'asof_daily_done': asof_daily_done,
        'asof_daily_planned': asof_daily_planned,
        'asof_daily_pct': _v73_pct(asof_daily_done, asof_daily_planned),
        'asof_weekly_done': asof_weekly_done,
        'asof_weekly_planned': asof_weekly_planned,
        'asof_weekly_pct': _v73_pct(asof_weekly_done, asof_weekly_planned),
        'overall_daily_done': overall_daily_done,
        'overall_daily_planned': overall_daily_planned,
        'overall_daily_pct': _v73_pct(overall_daily_done, overall_daily_planned),
        'overall_weekly_done': overall_weekly_done,
        'overall_weekly_planned': overall_weekly_planned,
        'overall_weekly_pct': _v73_pct(overall_weekly_done, overall_weekly_planned),

        'main_project': str(main_project or ''),
        'scenario': str(scenario or ''),
    }
